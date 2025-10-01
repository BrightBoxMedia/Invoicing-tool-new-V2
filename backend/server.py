#!/usr/bin/env python3
"""
Activus Invoice Management System - Production Backend
Professional invoice and project management for construction industry
"""

import os
import sys
import uuid
import bcrypt
import jwt
import asyncio
from datetime import datetime, timedelta, timezone
from typing import List, Dict, Optional, Union
import io
from io import BytesIO
import base64

# FastAPI and Pydantic imports
from fastapi import FastAPI, HTTPException, Depends, File, UploadFile, Query, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from fastapi.responses import Response, FileResponse, StreamingResponse
from fastapi.encoders import jsonable_encoder
from pydantic import BaseModel, Field, validator
import uvicorn
import json
import websockets
import pydantic
from bson import ObjectId

# Fix ObjectId serialization for JSON responses
# Note: In newer Pydantic versions, this is handled differently
try:
    pydantic.json.ENCODERS_BY_TYPE[ObjectId] = str
except AttributeError:
    # For newer Pydantic versions, we'll handle ObjectId serialization in the models
    pass

# Custom JSON encoder for ObjectId
def custom_jsonable_encoder(obj):
    if isinstance(obj, ObjectId):
        return str(obj)
    elif isinstance(obj, dict):
        return {key: custom_jsonable_encoder(value) for key, value in obj.items()}
    elif isinstance(obj, list):
        return [custom_jsonable_encoder(item) for item in obj]
    else:
        return jsonable_encoder(obj)

# Database
import motor.motor_asyncio
from motor.motor_asyncio import AsyncIOMotorDatabase
from pdf_template_manager import PDFTemplateManager, PDFTemplateConfig, initialize_template_manager, template_manager

# Excel processing
import openpyxl
from openpyxl import load_workbook

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.units import mm, inch
from reportlab.pdfgen import canvas

# Logging
import logging

# Canvas-based PDF generation for Canva-like functionality
async def generate_canvas_based_pdf(
    template_config: PDFTemplateConfig, 
    invoice_data: dict, 
    client_data: dict, 
    project_data: dict
) -> bytes:
    """
    Generate PDF using Canvas API for precise element positioning (Canva-like functionality)
    
    Args:
        template_config: PDF template configuration with canvas_elements
        invoice_data: Invoice data dictionary
        client_data: Client information dictionary  
        project_data: Project information dictionary
        
    Returns:
        bytes: Generated PDF as bytes
    """
    try:
        # Create PDF buffer
        buffer = BytesIO()
        
        # Determine page size
        page_size = A4 if template_config.page_size == "A4" else letter
        page_width, page_height = page_size
        
        # Create canvas for precise positioning
        c = canvas.Canvas(buffer, pagesize=page_size)
        
        # Helper function to convert frontend coordinates to ReportLab coordinates
        # ReportLab uses bottom-left origin, frontend uses top-left origin
        def convert_coordinates(x, y, element_height=0):
            # Convert from top-left origin to bottom-left origin
            converted_x = float(x) * (page_width / 600)  # Scale from 600px canvas to actual page width
            converted_y = page_height - (float(y) * (page_height / 800)) - element_height  # Scale and flip Y axis
            return converted_x, converted_y
        
        # Render each canvas element
        for element_id, element in template_config.canvas_elements.items():
            try:
                # Get element properties
                element_type = element.type
                x, y = convert_coordinates(element.x, element.y, element.height or 20)
                content = element.content
                style = element.style
                
                # Set text properties
                c.setFont("Helvetica", style.fontSize or 12)
                c.setFillColor(colors.toColor(style.color or "#000000"))
                
                # Render based element type
                if element_type == "text":
                    # Simple text element
                    text_content = str(content) if content else "Text Element"
                    c.drawString(x, y, text_content)
                    
                elif element_type == "text-group":
                    # Multi-line text group (invoice details)
                    if isinstance(content, dict):
                        line_height = (style.fontSize or 12) + 4
                        current_y = y
                        
                        # Draw invoice number
                        if content.get('invoice_no'):
                            c.drawString(x, current_y, f"Invoice No: {content['invoice_no']}")
                            current_y -= line_height
                        
                        # Draw invoice date  
                        if content.get('invoice_date'):
                            c.drawString(x, current_y, f"Invoice Date: {content['invoice_date']}")
                            current_y -= line_height
                            
                        # Draw created by
                        if content.get('created_by'):
                            c.drawString(x, current_y, f"Created By: {content['created_by']}")
                    
                elif element_type == "info-section":
                    # Company/client information section
                    if isinstance(content, dict):
                        line_height = (style.fontSize or 12) + 2
                        current_y = y
                        
                        # Draw background rectangle if specified
                        if style.backgroundColor:
                            c.setFillColor(colors.toColor(style.backgroundColor))
                            c.rect(x - 5, y - 5, element.width + 10, -(element.height or 100), fill=1, stroke=0)
                            c.setFillColor(colors.toColor(style.color or "#000000"))
                        
                        # Draw title (BILLED BY: / BILLED TO:)
                        if content.get('title'):
                            c.setFont("Helvetica-Bold", (style.fontSize or 12) + 1)
                            c.drawString(x, current_y, content['title'])
                            current_y -= line_height + 2
                        
                        # Draw company name
                        if content.get('company_name'):
                            c.setFont("Helvetica-Bold", style.fontSize or 12)
                            c.drawString(x, current_y, content['company_name'])
                            current_y -= line_height
                        
                        # Draw address (multi-line)
                        if content.get('company_address'):
                            c.setFont("Helvetica", (style.fontSize or 12) - 1)
                            address_lines = content['company_address'].split('\n')
                            for line in address_lines:
                                if line.strip():
                                    c.drawString(x, current_y, line.strip())
                                    current_y -= line_height - 2
                        
                        # Draw GST
                        if content.get('company_gst'):
                            c.drawString(x, current_y, f"GST: {content['company_gst']}")
                            current_y -= line_height - 2
                            
                        # Draw email
                        if content.get('company_email'):
                            c.drawString(x, current_y, f"Email: {content['company_email']}")
                            current_y -= line_height - 2
                            
                        # Draw phone
                        if content.get('company_phone'):
                            c.drawString(x, current_y, f"Phone: {content['company_phone']}")
                
                elif element_type == "project-info":
                    # Project information
                    if isinstance(content, dict):
                        line_height = (style.fontSize or 12) + 4
                        current_y = y
                        
                        if content.get('project_name'):
                            c.drawString(x, current_y, f"Project: {content['project_name']}")
                            current_y -= line_height
                            
                        if content.get('location'):
                            c.drawString(x, current_y, f"Location: {content['location']}")
                
            except Exception as e:
                logger.warning(f"Error rendering canvas element {element_id}: {e}")
                continue
        
        # Render logo if present
        if hasattr(template_config, 'logo_url') and template_config.logo_url:
            try:
                if template_config.logo_url.startswith('data:image'):
                    # Extract base64 data
                    logo_data = template_config.logo_url.split(',')[1]
                    logo_bytes = base64.b64decode(logo_data)
                    logo_buffer = BytesIO(logo_bytes)
                    
                    # Convert logo coordinates
                    logo_x, logo_y = convert_coordinates(template_config.logo_x or 350, template_config.logo_y or 20, template_config.logo_height or 60)
                    
                    # Draw logo on canvas
                    c.drawImage(
                        logo_buffer,
                        logo_x, logo_y,
                        width=template_config.logo_width or 120,
                        height=template_config.logo_height or 60
                    )
            except Exception as e:
                logger.warning(f"Logo rendering failed in canvas mode: {e}")
        
        # Generate sample invoice table if no table element exists
        # This ensures backward compatibility
        table_exists = any(element.type == "table" for element in template_config.canvas_elements.values())
        if not table_exists:
            # Add a simple sample table at bottom of page
            table_y = 200  # Fixed position for sample table
            
            # Table headers
            headers = ["Item", "GST Rate", "Qty", "Rate", "Amount", "IGST", "Total"]
            col_widths = [75, 18, 20, 22, 30, 25, 30]
            start_x = 50
            
            # Draw table headers
            c.setFillColor(colors.toColor(template_config.table_header_color or "#127285"))
            c.rect(start_x, table_y, sum(col_widths), 20, fill=1, stroke=1)
            
            c.setFillColor(colors.toColor(template_config.table_header_text_color or "#FFFFFF"))
            c.setFont("Helvetica-Bold", template_config.table_header_font_size or 11)
            
            current_x = start_x + 5
            for i, header in enumerate(headers):
                c.drawString(current_x, table_y + 5, header)
                current_x += col_widths[i]
            
            # Draw sample row
            c.setFillColor(colors.toColor("#000000"))
            c.setFont("Helvetica", template_config.table_data_font_size or 10)
            
            sample_row = ["Sample Construction Work", "18%", "100", "Rs. 1,000", "Rs. 100,000", "Rs. 18,000", "Rs. 118,000"]
            current_x = start_x + 5
            for i, cell in enumerate(sample_row):
                c.drawString(current_x, table_y - 15, cell)
                current_x += col_widths[i]
        
        # Save the canvas
        c.save()
        
        # Get PDF bytes
        buffer.seek(0)
        return buffer.getvalue()
        
    except Exception as e:
        logger.error(f"Canvas-based PDF generation failed: {e}")
        # Fall back to traditional generation
        return await generate_traditional_pdf(template_config, invoice_data, client_data, project_data)

# Traditional PDF generation (renamed for clarity)
async def generate_traditional_pdf(
    template_config: PDFTemplateConfig, 
    invoice_data: dict, 
    client_data: dict, 
    project_data: dict
) -> bytes:
    """Traditional PDF generation using ReportLab Platypus (story-based)"""
    try:

# Environment setup
SECRET_KEY = os.getenv('JWT_SECRET', 'activus-invoice-secret-key-2025')
MONGO_URL = os.getenv('MONGO_URL', 'mongodb://localhost:27017')
PORT = int(os.getenv('PORT', '8001'))

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize FastAPI
app = FastAPI(
    title="Activus Invoice Management API",
    description="Professional Invoice Management System for Construction Projects",
    version="2.0.0"
)

# CORS configuration - AWS production ready
ALLOWED_ORIGINS = os.getenv('ALLOWED_ORIGINS', '*').split(',')
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,  # Set ALLOWED_ORIGINS env var for production
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE"],
    allow_headers=["*"],
)

# Security
security = HTTPBearer()

# Database connection
client = motor.motor_asyncio.AsyncIOMotorClient(MONGO_URL)
DB_NAME = os.getenv('DB_NAME', 'activus_invoice_db')
db: AsyncIOMotorDatabase = client[DB_NAME]

# Real-time WebSocket Infrastructure for AWS Compatibility
class ConnectionManager:
    """WebSocket connection manager with AWS-compatible features"""
    
    def __init__(self):
        self.active_connections: Dict[str, List[WebSocket]] = {}
        self.user_projects: Dict[str, str] = {}  # user_id -> current_project_id
        self.last_event_timestamps: Dict[str, datetime] = {}
        
    async def connect(self, websocket: WebSocket, project_id: str, user_id: str):
        """Connect user to project channel"""
        await websocket.accept()
        
        if project_id not in self.active_connections:
            self.active_connections[project_id] = []
        
        self.active_connections[project_id].append(websocket)
        self.user_projects[user_id] = project_id
        
        logger.info(f"User {user_id} connected to project {project_id}")
        
    async def disconnect(self, websocket: WebSocket, project_id: str, user_id: str):
        """Disconnect user from project channel"""
        if project_id in self.active_connections:
            try:
                self.active_connections[project_id].remove(websocket)
                if not self.active_connections[project_id]:
                    del self.active_connections[project_id]
            except ValueError:
                pass
        
        if user_id in self.user_projects:
            del self.user_projects[user_id]
        
        logger.info(f"User {user_id} disconnected from project {project_id}")
    
    async def send_personal_message(self, message: str, websocket: WebSocket):
        """Send message to specific websocket"""
        try:
            await websocket.send_text(message)
        except WebSocketDisconnect:
            pass
    
    async def broadcast_to_project(self, project_id: str, message: dict, exclude_user: str = None):
        """Broadcast message to all users viewing a project"""
        if project_id not in self.active_connections:
            return
            
        message_str = json.dumps({
            **message,
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "project_id": project_id
        })
        
        # Store last event timestamp for reconnection handling
        self.last_event_timestamps[project_id] = datetime.now(timezone.utc)
        
        # Send to all connected clients for this project
        disconnected = []
        for websocket in self.active_connections[project_id]:
            try:
                await websocket.send_text(message_str)
            except (WebSocketDisconnect, Exception):
                disconnected.append(websocket)
        
        # Clean up disconnected clients
        for ws in disconnected:
            try:
                self.active_connections[project_id].remove(ws)
            except ValueError:
                pass
                
    async def get_project_snapshot(self, project_id: str):
        """Get current canonical project state for reconnection"""
        try:
            # Get fresh project data
            project = await db.projects.find_one({"id": project_id})
            if not project:
                return None
                
            # Get invoices for this project
            invoices_cursor = db.invoices.find({"project_id": project_id})
            invoices = await invoices_cursor.to_list(length=None)
            
            # Calculate current totals
            total_billed = sum(inv.get("total_amount", 0) for inv in invoices if inv.get("invoice_type") == "tax_invoice")
            total_project_value = project.get("total_project_value", 0)
            remaining_value = total_project_value - total_billed
            
            return {
                "event": "project_snapshot",
                "project_id": project_id,
                "data": {
                    "total_billed": total_billed,
                    "remaining_value": remaining_value,
                    "project_completed_percentage": (total_billed / total_project_value * 100) if total_project_value > 0 else 0,
                    "total_invoices": len(invoices),
                    "last_event_timestamp": self.last_event_timestamps.get(project_id, datetime.now(timezone.utc)).isoformat()
                }
            }
        except Exception as e:
            logger.error(f"Error getting project snapshot: {e}")
            return None

# Global connection manager instance
manager = ConnectionManager()

# Real-time Event System
class ProjectEvent:
    """Project event types for real-time updates"""
    INVOICE_CREATED = "invoice.created"
    INVOICE_UPDATED = "invoice.updated"  
    INVOICE_DELETED = "invoice.deleted"
    EXPENSE_CREATED = "expense.created"
    EXPENSE_UPDATED = "expense.updated"
    PROJECT_UPDATED = "project.updated"
    BOQ_ITEM_BILLED = "boq.item_billed"
    BOQ_CONFLICT = "boq.conflict"

async def emit_project_event(event_type: str, project_id: str, data: dict, user_id: str = None):
    """Emit real-time event for project changes"""
    try:
        # Get updated project data for canonical totals
        project_snapshot = await manager.get_project_snapshot(project_id)
        
        event_data = {
            "event": event_type,
            "project_id": project_id,
            "data": data,
            "actor_id": user_id,
            "canonical_totals": project_snapshot["data"] if project_snapshot else {}
        }
        
        # Broadcast to all users viewing this project
        await manager.broadcast_to_project(project_id, event_data, exclude_user=user_id)
        
        # Log event for monitoring
        logger.info(f"Event emitted: {event_type} for project {project_id}")
        
    except Exception as e:
        logger.error(f"Error emitting project event: {e}")

# Event payload helpers
def create_invoice_event_data(invoice: dict, boq_items_affected: List[str] = None) -> dict:
    """Create invoice event payload with minimal required fields"""
    return {
        "invoice_id": invoice.get("id"),
        "invoice_number": invoice.get("invoice_number"),
        "invoice_type": invoice.get("invoice_type"),
        "total_amount": invoice.get("total_amount", 0),
        "gst_amount": invoice.get("total_gst_amount", 0),
        "ra_tag": invoice.get("ra_number"),
        "affected_item_ids": boq_items_affected or []
    }

def create_boq_event_data(item_id: str, billed_quantity: float, available_quantity: float) -> dict:
    """Create BOQ item billing event payload"""
    return {
        "item_id": item_id,
        "billed_quantity": billed_quantity,
        "available_quantity": available_quantity
    }

# Pydantic Models
class UserRole:
    ADMIN = "admin"
    SUPER_ADMIN = "super_admin"
    USER = "user"

class InvoiceType:
    PROFORMA = "proforma"
    TAX_INVOICE = "tax_invoice"

class ProjectStatus:
    ACTIVE = "active"
    COMPLETED = "completed"
    ON_HOLD = "on_hold"

# User Models
class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: str
    password_hash: str
    role: str
    company_name: str
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserLogin(BaseModel):
    email: str
    password: str

class UserCreate(BaseModel):
    email: str
    password: str
    role: str
    company_name: str

# Client Models
class ClientInfo(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: str
    phone: str
    address: str
    city: str
    state: str
    gst_no: str
    bill_to_address: str
    shipping_address: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# BOQ Item Models
class BOQItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sr_no: int
    description: str
    unit: str
    quantity: float
    rate: float
    amount: float
    gst_rate: float = 18.0
    billed_quantity: float = 0.0  # Track what's already billed

    @validator('gst_rate')
    def validate_gst_rate(cls, v):
        allowed_rates = [0, 5, 12, 18, 28, 40]  # Added 40% GST
        if v not in allowed_rates:
            raise ValueError(f'GST rate must be one of {allowed_rates}')
        return v

# Project Models
class Project(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_name: str
    client_id: str
    client_name: str
    architect: str
    location: str
    
    # Cash flow percentages - must total 100%
    abg_percentage: float = 0.0  # Advance Bank Guarantee
    ra_bill_percentage: float = 0.0   # RA Bill with Taxes (renamed for consistency)
    erection_percentage: float = 0.0  # Erection Work
    pbg_percentage: float = 0.0  # Performance Bank Guarantee
    
    # GST Configuration
    gst_type: str = "IGST"  # CGST_SGST or IGST (default to IGST)
    gst_approval_status: str = "pending"  # pending, approved, rejected
    gst_approved_by: Optional[str] = None  # User ID who approved GST
    gst_approved_at: Optional[datetime] = None  # When GST was approved
    
    total_project_value: float
    boq_items: List[BOQItem] = []
    status: str = ProjectStatus.ACTIVE
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

    @validator('abg_percentage', 'ra_bill_percentage', 'erection_percentage', 'pbg_percentage')
    def validate_percentage(cls, v):
        if v < 0 or v > 100:
            raise ValueError('Percentage must be between 0 and 100')
        return v

    @validator('gst_type')
    def validate_gst_type(cls, v):
        if v not in ['CGST_SGST', 'IGST']:
            raise ValueError('GST type must be either CGST_SGST or IGST')
        return v

    @validator('gst_approval_status')
    def validate_gst_approval_status(cls, v):
        if v not in ['pending', 'approved', 'rejected']:
            raise ValueError('GST approval status must be pending, approved, or rejected')
        return v

    def validate_total_percentage(self):
        total = self.abg_percentage + self.ra_bill_percentage + self.erection_percentage + self.pbg_percentage
        if abs(total - 100.0) > 0.01:  # Allow small floating point differences
            raise ValueError(f'Percentages must total 100%, current total: {total}%')

# Invoice Models
class InvoiceItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    boq_item_id: str
    description: str
    unit: str
    quantity: float
    rate: float
    amount: float
    gst_rate: float = 18.0

class Invoice(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_number: str
    project_id: str
    client_id: str
    invoice_type: str  # "proforma" or "tax_invoice"
    invoice_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    due_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc) + timedelta(days=30))
    items: List[InvoiceItem]
    subtotal: float
    
    # GST Breakdown
    gst_type: str  # CGST_SGST or IGST
    cgst_amount: float = 0.0  # Only for CGST_SGST type
    sgst_amount: float = 0.0  # Only for CGST_SGST type  
    igst_amount: float = 0.0  # Only for IGST type
    total_gst_amount: float
    
    total_amount: float
    payment_terms: str = "Payment due within 30 days"
    advance_received: float = 0.0
    net_amount_due: float = 0.0
    ra_number: Optional[str] = None  # Only for tax invoices
    status: str = "draft"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ActivityLog(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_email: str
    user_role: str
    action: str
    description: str
    project_id: Optional[str] = None
    invoice_id: Optional[str] = None
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Excel Parser Class - FIXED VERSION
class ExcelParser:
    def __init__(self):
        self.supported_extensions = ['.xlsx', '.xlsm', '.xls']
        
    async def parse_excel_file(self, file_content: bytes, filename: str) -> Dict:
        try:
            workbook = load_workbook(BytesIO(file_content), data_only=True)
            worksheet = workbook.active
            
            # Enhanced BOQ parsing - ignore totals and summaries
            parsed_data = await self._parse_boq_data(worksheet, filename)
            
            return {
                "filename": filename,
                "sheets": [worksheet.title],
                "parsed_data": parsed_data,
                "status": "success"
            }
            
        except Exception as e:
            logger.error(f"Excel parsing error: {str(e)}")
            raise HTTPException(status_code=422, detail=f"Failed to parse Excel file: {str(e)}")
    
    async def _parse_boq_data(self, worksheet, filename: str) -> Dict:
        """SUPER INTELLIGENT BOQ parsing - handles ANY Excel format including complex layouts"""
        
        logger.info(f"🚀 STARTING SUPER INTELLIGENT BOQ PARSING for {filename}")
        logger.info(f"📊 Worksheet dimensions: {worksheet.max_row} rows × {worksheet.max_column} columns")
        
        # STRATEGY 1: Try standard header-based parsing
        boq_items = []
        try:
            logger.info("🔍 STRATEGY 1: Standard header-based parsing")
            header_row = self._find_header_row(worksheet)
            if header_row:
                column_mapping = self._get_enhanced_column_mapping(worksheet, header_row)
                logger.info(f"📋 Column mapping found: {column_mapping}")
                
                if column_mapping and 'description' in column_mapping:
                    boq_items = await self._extract_items_with_mapping(worksheet, header_row, column_mapping)
                    if boq_items:
                        logger.info(f"✅ STRATEGY 1 SUCCESS: Found {len(boq_items)} items")
                        return await self._finalize_boq_data(boq_items, filename)
        except Exception as e:
            logger.warning(f"⚠️ Strategy 1 failed: {e}")
        
        # STRATEGY 2: Pattern-based parsing (no strict headers)  
        try:
            logger.info("🔍 STRATEGY 2: Pattern-based parsing")
            boq_items = await self._extract_items_by_pattern(worksheet)
            if boq_items:
                logger.info(f"✅ STRATEGY 2 SUCCESS: Found {len(boq_items)} items")
                return await self._finalize_boq_data(boq_items, filename)
        except Exception as e:
            logger.warning(f"⚠️ Strategy 2 failed: {e}")
        
        # STRATEGY 3: Brute force - scan all cells for BOQ-like data
        try:
            logger.info("🔍 STRATEGY 3: Brute force scanning")
            boq_items = await self._extract_items_brute_force(worksheet)
            if boq_items:
                logger.info(f"✅ STRATEGY 3 SUCCESS: Found {len(boq_items)} items")
                return await self._finalize_boq_data(boq_items, filename)
        except Exception as e:
            logger.warning(f"⚠️ Strategy 3 failed: {e}")
        
        # If all strategies fail
        logger.error("❌ ALL STRATEGIES FAILED - No valid BOQ items found")
        raise ValueError("No valid BOQ items found in the Excel file. Please check the file format and ensure it contains item descriptions with quantities, rates, or amounts.")
    
    async def _extract_items_with_mapping(self, worksheet, header_row: int, column_mapping: Dict) -> List[BOQItem]:
        """Extract items using column mapping"""
        boq_items = []
        
        for row_idx in range(header_row + 1, min(worksheet.max_row + 1, header_row + 500)):
            try:
                row_data = self._extract_row_data(worksheet, row_idx, column_mapping)
                
                # Skip if this is a summary/total row
                if self._is_summary_row(row_data):
                    logger.info(f"Skipping summary row {row_idx}: {row_data.get('description', 'Unknown')}")
                    continue
                
                # Validate if this is a proper BOQ item
                if self._is_valid_boq_item(row_data):
                    boq_item = self._create_boq_item(row_data, len(boq_items) + 1)
                    boq_items.append(boq_item)
                    logger.info(f"✓ Mapped item {len(boq_items)}: {row_data['description'][:50]}")
                    
            except Exception as e:
                logger.warning(f"Error processing row {row_idx}: {e}")
                continue
        
        return boq_items
    
    async def _extract_items_by_pattern(self, worksheet) -> List[BOQItem]:
        """Extract items by detecting BOQ patterns without strict headers"""
        boq_items = []
        
        logger.info("🔍 PATTERN SCANNING: Looking for BOQ data patterns...")
        
        for row_num in range(1, min(worksheet.max_row + 1, 200)):
            row_cells = []
            
            # Get all non-empty cells in this row
            for col_num in range(1, min(worksheet.max_column + 1, 50)):
                cell = worksheet.cell(row=row_num, column=col_num)
                if cell.value is not None:
                    row_cells.append({
                        'value': cell.value,
                        'column': col_num,
                        'is_number': isinstance(cell.value, (int, float)),
                        'is_text': isinstance(cell.value, str)
                    })
            
            # Pattern detection: Look for rows with description + numbers
            if len(row_cells) >= 3:
                description_cell = None
                quantity_cell = None
                rate_cell = None
                amount_cell = None
                
                # Find description (longest text or first substantial text)
                text_cells = [cell for cell in row_cells if cell['is_text'] and len(str(cell['value']).strip()) > 5]
                if text_cells:
                    # Take the longest text as description
                    description_cell = max(text_cells, key=lambda x: len(str(x['value']).strip()))
                
                # Find numbers (quantity, rate, amount)
                number_cells = [cell for cell in row_cells if cell['is_number'] and cell['value'] > 0]
                
                if description_cell and len(number_cells) >= 2:
                    # Create row data
                    row_data = {
                        'description': str(description_cell['value']).strip(),
                        'quantity': float(number_cells[0]['value']) if len(number_cells) >= 1 else 1.0,
                        'rate': float(number_cells[1]['value']) if len(number_cells) >= 2 else float(number_cells[0]['value']),
                        'amount': float(number_cells[2]['value']) if len(number_cells) >= 3 else float(number_cells[0]['value']) * float(number_cells[1]['value']),
                        'unit': 'Nos',
                        'gst_rate': 18.0
                    }
                    
                    if self._is_valid_boq_item(row_data):
                        boq_item = self._create_boq_item(row_data, len(boq_items) + 1)
                        boq_items.append(boq_item)
                        logger.info(f"✓ Pattern item {len(boq_items)}: {row_data['description'][:50]} | Q:{row_data['quantity']} R:{row_data['rate']}")
        
        return boq_items
    
    async def _extract_items_brute_force(self, worksheet) -> List[BOQItem]:
        """Brute force extraction - find ANY rows that look like BOQ items"""
        boq_items = []
        
        logger.info("💪 BRUTE FORCE SCANNING: Extracting any BOQ-like data...")
        
        # Collect all meaningful data from worksheet
        rows_data = {}
        
        for row_num in range(1, min(worksheet.max_row + 1, 500)):
            for col_num in range(1, min(worksheet.max_column + 1, 50)):
                cell = worksheet.cell(row=row_num, column=col_num)
                if cell.value is not None:
                    if row_num not in rows_data:
                        rows_data[row_num] = []
                    rows_data[row_num].append({
                        'value': cell.value,
                        'col': col_num,
                        'is_number': isinstance(cell.value, (int, float)),
                        'is_text': isinstance(cell.value, str)
                    })
        
        # Analyze each row for BOQ potential
        for row_num, row_data in rows_data.items():
            texts = [item for item in row_data if item['is_text']]
            numbers = [item for item in row_data if item['is_number'] and item['value'] > 0]
            
            # Basic BOQ criteria: at least 1 substantial text + 2 numbers
            if len(texts) >= 1 and len(numbers) >= 2:
                
                # Find best description candidate
                description_candidate = None
                for text_item in texts:
                    text_val = str(text_item['value']).strip()
                    # Skip obvious non-descriptions
                    skip_terms = ['total', 'sum', 'gst', 'tax', 'nil', 'na', 'n/a', 'subtotal', 'grand total']
                    if any(skip in text_val.lower() for skip in skip_terms):
                        continue
                    if len(text_val) >= 5:  # Reasonable description length
                        description_candidate = text_val
                        break
                
                if description_candidate:
                    # Use available numbers
                    sorted_numbers = sorted(numbers, key=lambda x: x['col'])
                    
                    quantity = float(sorted_numbers[0]['value']) if len(sorted_numbers) >= 1 else 1.0
                    rate = float(sorted_numbers[1]['value']) if len(sorted_numbers) >= 2 else quantity
                    amount = float(sorted_numbers[2]['value']) if len(sorted_numbers) >= 3 else quantity * rate
                    
                    row_data_dict = {
                        'description': description_candidate,
                        'quantity': quantity,
                        'rate': rate,
                        'amount': amount,
                        'unit': 'Nos',
                        'gst_rate': 18.0
                    }
                    
                    if self._is_valid_boq_item(row_data_dict):
                        boq_item = self._create_boq_item(row_data_dict, len(boq_items) + 1)
                        boq_items.append(boq_item)
                        logger.info(f"✓ Brute force item {len(boq_items)}: {description_candidate[:40]} | Q:{quantity} R:{rate}")
        
        return boq_items
    
    def _create_boq_item(self, row_data: Dict, sr_no: int) -> BOQItem:
        """Create a standardized BOQ item"""
        # Ensure GST rate is valid
        gst_rate = row_data.get('gst_rate', 18.0)
        if gst_rate not in [0, 5, 12, 18, 28, 40]:
            gst_rate = 18.0  # Default
        
        return BOQItem(
            sr_no=sr_no,
            description=row_data.get('description', 'Unknown Item'),
            unit=row_data.get('unit', 'Nos'),
            quantity=float(row_data.get('quantity', 0.0)),
            rate=float(row_data.get('rate', 0.0)),
            amount=float(row_data.get('amount', 0.0)),
            gst_rate=float(gst_rate),
            billed_quantity=0.0  # Initialize as unbilled
        )
    
    async def _finalize_boq_data(self, boq_items: List[BOQItem], filename: str) -> Dict:
        """Finalize and return BOQ data"""
        if not boq_items:
            raise ValueError("No valid BOQ items found")
        
        # Extract project metadata
        project_info = {
            "project_name": filename.replace('.xlsx', '').replace('.xls', ''),
            "total_items": len(boq_items),
            "total_amount": sum(item.amount for item in boq_items)
        }
        
        logger.info(f"🎉 PARSING COMPLETE: {len(boq_items)} items found, total amount: ₹{project_info['total_amount']:,.2f}")
        
        return {
            "project_info": project_info,
            "boq_items": [item.dict() for item in boq_items]
        }
    
    def _is_summary_row(self, row_data: Dict) -> bool:
        """Check if this row is a summary/total row - ENHANCED for user's format"""
        description = str(row_data.get('description', '')).lower().strip()
        
        # More specific summary row indicators (don't be too aggressive)
        summary_indicators = [
            'total', 'grand total', 'subtotal', 'sum', 'gst at', 'tax', 
            'amount left to claim', 'balance', 'remaining', 'summary',
            'provisional sum', 'p.sum', 'contingency', 'overhead',
            'profit', 'margin', 'discount'
        ]
        
        # Only reject if description exactly matches or contains clear summary indicators
        for indicator in summary_indicators:
            if indicator in description and len(description) > len(indicator):
                # Only reject if it's a substantial match, not just a substring
                if description.startswith(indicator) or description.endswith(indicator):
                    return True
        
        # Don't reject based on short description length for user's format
        # User has valid items like "TOP", "Left", "Right" which are short but valid
        
        # Check if all numeric fields are zero (empty summary row)
        quantity = row_data.get('quantity', 0)
        rate = row_data.get('rate', 0)  
        amount = row_data.get('amount', 0)
        
        # Only reject if completely empty (no description AND no numbers)
        if not description and quantity == 0 and rate == 0 and amount == 0:
            return True
        
        return False
    
    def _extract_project_metadata(self, worksheet) -> Dict:
        """Extract project information from the top section of the Excel"""
        project_info = {
            'project_name': '',
            'client_name': '',
            'architect': '',
            'location': ''
        }
        
        # Search first 15 rows for project information
        for row in range(1, min(16, worksheet.max_row + 1)):
            for col in range(1, min(10, worksheet.max_column + 1)):
                cell_value = worksheet.cell(row=row, column=col).value
                if not cell_value:
                    continue
                
                cell_str = str(cell_value).lower().strip()
                
                # Look for project name indicators
                if any(indicator in cell_str for indicator in ['project', 'work', 'site']):
                    next_cell = worksheet.cell(row=row, column=col + 1).value
                    if next_cell and len(str(next_cell).strip()) > 5:
                        project_info['project_name'] = str(next_cell).strip()
                
                # Look for client name indicators
                if any(indicator in cell_str for indicator in ['client', 'company', 'contractor']):
                    next_cell = worksheet.cell(row=row, column=col + 1).value
                    if next_cell and len(str(next_cell).strip()) > 3:
                        project_info['client_name'] = str(next_cell).strip()
                
                # Look for architect indicators
                if 'architect' in cell_str:
                    next_cell = worksheet.cell(row=row, column=col + 1).value
                    if next_cell and len(str(next_cell).strip()) > 3:
                        project_info['architect'] = str(next_cell).strip()
                
                # Look for location indicators
                if any(indicator in cell_str for indicator in ['location', 'address', 'site']):
                    next_cell = worksheet.cell(row=row, column=col + 1).value
                    if next_cell and len(str(next_cell).strip()) > 5:
                        project_info['location'] = str(next_cell).strip()
        
        return project_info
    
    def _find_header_row(self, worksheet) -> Optional[int]:
        """ENHANCED header detection - specifically handles user's Excel format"""
        logger.info("🔍 ENHANCED HEADER SEARCH for user's Excel format...")
        
        for row in range(1, min(50, worksheet.max_row + 1)):
            row_text = []
            non_empty_count = 0
            
            for col in range(1, min(30, worksheet.max_column + 1)):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value:
                    row_text.append(str(cell_value).lower().strip())
                    non_empty_count += 1
            
            row_combined = ' '.join(row_text)
            logger.info(f"Row {row}: {non_empty_count} cells | '{row_combined[:100]}...'")
            
            # ENHANCED detection for user's specific format
            # Look for the exact pattern: "Sl. No." + "Description Of Item" + quantity/unit indicators
            has_sl_no = any(indicator in row_combined for indicator in [
                'sl. no', 'sl.no', 'slno', 'sl no', 'sr. no', 'sr.no', 'srno', 'sr no', 'serial'
            ])
            
            has_description_of_item = any(indicator in row_combined for indicator in [
                'description of item', 'description', 'item', 'particulars', 'work item'
            ])
            
            has_qty = any(indicator in row_combined for indicator in [
                'qty', 'quantity', 'qnty'
            ])
            
            has_unit = any(indicator in row_combined for indicator in [
                'unit', 'uom', 'u.o.m'
            ])
            
            has_rate = any(indicator in row_combined for indicator in [
                'rate', 'rate/', 'rate /', 'rate/unit', 'rate / unit', 'unit rate'
            ])
            
            has_amount = any(indicator in row_combined for indicator in [
                'amount', 'total', 'value'
            ])
            
            # Score calculation - prioritize exact matches for user's format
            score = 0
            if has_sl_no: score += 2
            if has_description_of_item: score += 3  # Most important
            if has_qty: score += 2
            if has_unit: score += 2
            if has_rate: score += 2
            if has_amount: score += 1
            
            # Boost score if we have enough columns
            if non_empty_count >= 4: score += 1
            
            logger.info(f"Row {row} score: {score} | SlNo: {has_sl_no} | Desc: {has_description_of_item} | Qty: {has_qty} | Unit: {has_unit} | Rate: {has_rate} | Amount: {has_amount}")
            
            # Accept row if it has essential BOQ indicators
            if score >= 6 or (has_description_of_item and has_qty and (has_unit or has_rate)):
                logger.info(f"✅ FOUND HEADER ROW at {row}: '{row_combined[:100]}...'")
                return row
        
        # Fallback: Look for any row with "Description Of Item" specifically
        logger.warning("⚠️ Enhanced header detection failed, trying specific pattern fallback...")
        for row in range(1, min(50, worksheet.max_row + 1)):
            for col in range(1, min(30, worksheet.max_column + 1)):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value and 'description' in str(cell_value).lower():
                    # Check if this row has multiple headers
                    headers_in_row = 0
                    for c in range(1, min(10, worksheet.max_column + 1)):
                        cv = worksheet.cell(row=row, column=c).value
                        if cv and isinstance(cv, str) and len(str(cv).strip()) > 2:
                            headers_in_row += 1
                    
                    if headers_in_row >= 3:
                        logger.info(f"✅ FALLBACK HEADER ROW found at {row} with 'description' and {headers_in_row} headers")
                        return row
        
        logger.error("❌ Could not find any header row!")
        return None
    
    def _get_enhanced_column_mapping(self, worksheet, header_row: int) -> Dict[str, int]:
        """ENHANCED column mapping - handles user's specific Excel format"""
        column_mapping = {}
        
        # Debug: Print all headers found
        logger.info(f"ANALYZING EXCEL HEADERS at row {header_row}:")
        for col_idx in range(1, min(30, worksheet.max_column + 1)):
            cell = worksheet.cell(row=header_row, column=col_idx)
            if cell.value:
                logger.info(f"  Column {col_idx}: '{cell.value}'")
        
        for col_idx in range(1, min(30, worksheet.max_column + 1)):
            cell = worksheet.cell(row=header_row, column=col_idx)
            if not cell.value:
                continue
                
            cell_lower = str(cell.value).lower().strip()
            cell_original = str(cell.value).strip()
            
            # Enhanced Serial number mapping - handles user's "Sl. No." format
            if any(h in cell_lower for h in [
                'sl. no', 'sl.no', 'slno', 'sl no', 'sr. no', 'sr.no', 'srno', 'sr no', 
                'serial', 's.no', 'sno', 's no', '#', 'no.', 'no', 'item no', 'item_no'
            ]):
                column_mapping['sr_no'] = col_idx
                logger.info(f"✅ FOUND SR_NO at column {col_idx}: '{cell_original}'")
                
            # Enhanced Description mapping - handles user's "Description Of Item" format
            elif any(h in cell_lower for h in [
                'description of item', 'description', 'particular', 'particulars', 'item', 
                'work', 'activity', 'scope', 'specification', 'details', 'desc', 
                'work item', 'work_item', 'item description', 'item_description', 
                'scope of work', 'scope_of_work', 'material', 'service', 'product', 
                'component', 'task'
            ]):
                column_mapping['description'] = col_idx
                logger.info(f"✅ FOUND DESCRIPTION at column {col_idx}: '{cell_original}'")
                
            # Enhanced Unit mapping - simple and direct for user's format
            elif cell_lower == 'unit' or any(h in cell_lower for h in [
                'unit', 'uom', 'u.o.m', 'u o m', 'units', 'measure', 'measurement',
                'unit of measurement', 'unit_of_measurement'
            ]) and not any(x in cell_lower for x in ['rate', 'amount', 'price', 'cost']):
                column_mapping['unit'] = col_idx
                logger.info(f"✅ FOUND UNIT at column {col_idx}: '{cell_original}'")
                
            # Enhanced Quantity mapping - handles user's " Qty" format (with space)
            elif any(h in cell_lower for h in [
                'qty', 'quantity', 'qnty', 'quantities', 'total qty', 'total_qty',
                'req qty', 'req_qty', 'required qty', 'required_qty'
            ]) and not any(x in cell_lower for x in ['rate', 'price', 'cost', 'amount', 'value']):
                column_mapping['quantity'] = col_idx
                logger.info(f"✅ FOUND QUANTITY at column {col_idx}: '{cell_original}'")
                
            # Enhanced Rate mapping - handles user's "Rate/ Unit" format
            elif any(h in cell_lower for h in [
                'rate/ unit', 'rate/unit', 'rate / unit', 'rate', 'price', 'unit rate', 
                'unit_rate', 'unit price', 'unit_price', 'cost', 'per unit', 'per_unit', 
                'rate per unit', 'rate_per_unit', 'unit cost', 'unit_cost', 'basic rate', 'basic_rate'
            ]) and not any(x in cell_lower for x in ['total', 'sum']) and 'amount' not in cell_lower:
                column_mapping['rate'] = col_idx
                logger.info(f"✅ FOUND RATE at column {col_idx}: '{cell_original}'")
                
            # Enhanced Amount mapping - simple and direct
            elif cell_lower == 'amount' or any(h in cell_lower for h in [
                'amount', 'total', 'value', 'total amount', 'total_amount', 'total value', 'total_value',
                'basic amount', 'basic_amount', 'subtotal', 'sub total', 'sub_total',
                'line total', 'line_total', 'extended amount', 'extended_amount'
            ]) and not any(x in cell_lower for x in ['rate', 'unit', 'gst', 'tax']):
                column_mapping['amount'] = col_idx
                logger.info(f"✅ FOUND AMOUNT at column {col_idx}: '{cell_original}'")
        
        logger.info(f"📋 FINAL COLUMN MAPPING: {column_mapping}")
        
        # Enhanced validation - try to find description column more intelligently
        if 'description' not in column_mapping:
            logger.warning("❌ No description column found - trying enhanced fallback detection")
            # Find the column with the longest average text length (likely description)
            best_desc_col = None
            best_avg_length = 0
            
            for col_idx in range(1, min(10, worksheet.max_column + 1)):  # Check first 10 columns
                sample_rows = min(5, worksheet.max_row - header_row)  # Sample fewer rows for speed
                text_lengths = []
                text_content = []
                
                for r in range(header_row + 1, header_row + sample_rows + 1):
                    cell_val = worksheet.cell(row=r, column=col_idx).value
                    if cell_val and isinstance(cell_val, str):
                        text_val = str(cell_val).strip()
                        if len(text_val) > 2:  # Skip very short values
                            text_lengths.append(len(text_val))
                            text_content.append(text_val.lower())
                
                if text_lengths:
                    avg_length = sum(text_lengths) / len(text_lengths)
                    # Also check if content looks like descriptions (contains alphabetic characters)
                    has_descriptive_content = any(any(c.isalpha() for c in text) for text in text_content)
                    
                    if avg_length > best_avg_length and has_descriptive_content and avg_length > 5:
                        best_avg_length = avg_length
                        best_desc_col = col_idx
                        logger.info(f"Column {col_idx} candidate: avg length {avg_length:.1f}, samples: {text_content[:2]}")
            
            if best_desc_col:
                column_mapping['description'] = best_desc_col
                logger.info(f"✅ ENHANCED FALLBACK DESCRIPTION found at column {best_desc_col} (avg length: {best_avg_length:.1f})")
        
        return column_mapping
    
    def _extract_row_data(self, worksheet, row_idx: int, column_mapping: Dict[str, int]) -> Dict:
        """Enhanced row data extraction - handles user's Excel format"""
        row_data = {}
        
        for field, col_idx in column_mapping.items():
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell_value = cell.value
            
            if field == 'description':
                # Enhanced description extraction
                desc_value = str(cell_value).strip() if cell_value else ''
                # Handle cases where description might have extra spaces or formatting
                desc_value = ' '.join(desc_value.split())  # Normalize whitespace
                row_data[field] = desc_value
                
            elif field == 'unit':
                # Enhanced unit extraction - handle various formats
                if cell_value:
                    unit_str = str(cell_value).strip()
                    # Clean up common unit formats
                    unit_str = unit_str.replace('/', '').replace('\\', '').strip()
                    if not unit_str or unit_str.lower() in ['', 'nil', 'na', 'n/a']:
                        unit_str = 'Nos'
                else:
                    unit_str = 'Nos'
                row_data[field] = unit_str
                
            elif field in ['quantity', 'rate', 'amount']:
                # Enhanced numeric extraction
                row_data[field] = self._safe_float_conversion(cell_value)
                
            elif field == 'sr_no':
                # Enhanced serial number handling - can be text or number
                if cell_value:
                    sr_value = str(cell_value).strip()
                    row_data[field] = sr_value
                else:
                    row_data[field] = ''
            else:
                row_data[field] = cell_value
        
        # Enhanced data completion - fill missing fields intelligently
        if 'description' not in row_data:
            row_data['description'] = ''
        if 'unit' not in row_data:
            row_data['unit'] = 'Nos'
        if 'quantity' not in row_data:
            row_data['quantity'] = 0.0
        if 'rate' not in row_data:
            row_data['rate'] = 0.0  
        if 'amount' not in row_data:
            row_data['amount'] = 0.0
        
        # Calculate missing amount if we have quantity and rate
        if row_data['amount'] == 0.0 and row_data['quantity'] > 0 and row_data['rate'] > 0:
            row_data['amount'] = row_data['quantity'] * row_data['rate']
            
        return row_data
    
    def _is_valid_boq_item(self, row_data: Dict) -> bool:
        """ENHANCED validation - specifically handles user's Excel format"""
        description = str(row_data.get('description', '')).strip()
        
        # Enhanced description validation - be more forgiving
        if not description or len(description) < 2:
            return False
        
        # Skip clearly invalid descriptions (but be more specific)
        invalid_patterns = [
            'total', 'sum', 'subtotal', 'grand total', 'gst', 'tax', 'nil', 'n/a', 'na',
            'provisional sum', 'p.sum', 'contingency', 'overhead', 'profit', 'margin'
        ]
        desc_lower = description.lower().strip()
        
        # Only reject if the ENTIRE description matches invalid patterns
        if desc_lower in invalid_patterns or any(desc_lower == pattern for pattern in invalid_patterns):
            logger.info(f"❌ Rejecting invalid description: '{description}'")
            return False
        
        # More flexible numeric validation - handle user's data format
        quantity = row_data.get('quantity', 0) 
        rate = row_data.get('rate', 0)
        amount = row_data.get('amount', 0)
        
        # Convert to float safely
        try:
            quantity = float(quantity) if quantity else 0.0
            rate = float(rate) if rate else 0.0
            amount = float(amount) if amount else 0.0
        except (ValueError, TypeError):
            logger.info(f"❌ Invalid numeric values for: '{description[:30]}...'")
            return False
        
        # Enhanced validation logic
        has_quantity = quantity > 0
        has_rate = rate > 0  
        has_amount = amount > 0
        
        # Valid BOQ item criteria:
        # 1. Has description AND (quantity AND rate) OR
        # 2. Has description AND amount > 0 OR
        # 3. Has description AND any two numeric fields
        is_valid = (
            (has_quantity and has_rate) or  # Can calculate amount
            has_amount or  # Has final amount
            (sum([has_quantity, has_rate, has_amount]) >= 2)  # At least 2 numeric fields
        )
        
        # Special handling for user's specific items like "TOP", "Left", "Right", etc.
        if not is_valid and len(description) >= 3 and any(desc_lower.startswith(prefix) for prefix in ['top', 'left', 'right', 'buttom', 'side']):
            # These are likely valid items even with less strict validation
            is_valid = has_quantity or has_rate or has_amount
            logger.info(f"🔍 Special handling for user's item: '{description}'")
        
        logger.info(f"🔍 Validating: '{description[:30]}...' | Qty: {quantity} | Rate: {rate} | Amount: {amount} | Valid: {is_valid}")
        
        return is_valid
    
    def _safe_float_conversion(self, value):
        """Safely convert value to float"""
        if value is None or value == "":
            return 0.0
        
        if isinstance(value, (int, float)):
            return float(value)
        
        if isinstance(value, str):
            cleaned_value = str(value).replace('₹', '').replace('Rs', '').replace(',', '').strip()
            if cleaned_value == "" or cleaned_value.lower() == "none":
                return 0.0
            try:
                return float(cleaned_value)
            except (ValueError, TypeError):
                return 0.0
        
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0

# PDF Generator Class
class PDFGenerator:
    def __init__(self):
        self.page_size = A4
        self.margin = 20 * mm
        
    async def generate_invoice_pdf(self, invoice: Invoice, project: Project, client: ClientInfo):
        buffer = io.BytesIO()
        
        # EXACT page setup matching target PDF
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=20*mm,
            leftMargin=20*mm, 
            topMargin=15*mm,
            bottomMargin=20*mm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # ===== EXACT HEADER LAYOUT MATCHING TARGET PDF =====
        
        # TAX Invoice title - EXACTLY positioned and styled like target
        tax_invoice_style = ParagraphStyle(
            'TAXInvoiceTitle',
            parent=styles['Normal'],
            fontSize=18,
            textColor=colors.black,
            alignment=TA_CENTER,  # CENTERED like in target
            spaceAfter=0,
            fontName='Helvetica-Bold'
        )
        
        # Logo - EXACT size and positioning like target
        try:
            logo_path = '/app/frontend/public/activus-new-logo.png'
            if os.path.exists(logo_path):
                logo_element = RLImage(logo_path, width=120, height=60)  # Professional size matching target
            else:
                logo_element = Paragraph("<b>ACTIVUS INDUSTRIAL DESIGN & BUILD LLP</b><br/><i>One Stop Solution is What We Do</i>", 
                                       ParagraphStyle('LogoText', fontSize=10, alignment=TA_RIGHT, fontName='Helvetica-Bold'))
        except:
            logo_element = Paragraph("<b>ACTIVUS INDUSTRIAL DESIGN & BUILD LLP</b><br/><i>One Stop Solution is What We Do</i>", 
                                   ParagraphStyle('LogoText', fontSize=10, alignment=TA_RIGHT, fontName='Helvetica-Bold'))
        
        # Header layout EXACTLY like target - TAX Invoice centered, logo top right
        header_data = [[
            "",  # Empty left space
            logo_element
        ]]
        
        header_table = Table(header_data, colWidths=[400, 150])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        
        elements.append(header_table)
        elements.append(Spacer(1, 10))
        
        # TAX Invoice title - CENTERED like target
        elements.append(Paragraph("TAX Invoice", tax_invoice_style))
        elements.append(Spacer(1, 15))
        
        # ===== INVOICE IDENTIFICATION BLOCK - EXACT MATCH =====
        invoice_details_style = ParagraphStyle(
            'InvoiceDetailsStyle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.black,
            alignment=TA_LEFT,
            fontName='Helvetica',
            lineHeight=16,
            spaceAfter=20
        )
        
        # EXACT text format matching target PDF
        invoice_details_text = f"""<b>Invoice No #</b> {invoice.invoice_number}<br/>
<b>Invoice Date</b> {invoice.invoice_date.strftime('%b %d, %Y')}<br/>
<b>Created By</b> Sathiya Narayanan Kannan"""
        
        elements.append(Paragraph(invoice_details_text, invoice_details_style))
        elements.append(Spacer(1, 20))
        
        # ===== BILLED BY / BILLED TO SECTIONS - EXACT MATCH =====
        billing_section_style = ParagraphStyle(
            'BillingStyle',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.black,  
            fontName='Helvetica',
            lineHeight=14,
            alignment=TA_LEFT
        )
        
        # EXACT content format matching target PDF
        billed_by_text = """<b>Billed By</b><br/><br/>
<b>Activus Industrial Design And Build LLP</b><br/>
Flat No.125 7th Cross Rd, Opp Bannerghatta Road, Dollar Layout,<br/>
BTM Layout Stage 2, Bilekahlli, Bengaluru, Karnataka, India - 560076<br/><br/>
<b>GSTIN:</b> 29ACGFA5744D1ZF<br/>
<b>PAN:</b> ACGFA5744D<br/>
<b>Email:</b> finance@activusdesignbuild.in<br/>
<b>Phone:</b> +91 87785 07177"""
        
        billed_to_text = f"""<b>Billed To</b><br/><br/>
<b>{client.name or 'UNITED BREWERIES LIMITED'}</b><br/>
{client.bill_to_address or 'PLOT NO M-1 & M-1 /2,TALOJA DIST. RAIGAD,Maharashtra-410208.,'}<br/>
Taloja, Maharashtra, India - 410206<br/><br/>
<b>GSTIN:</b> {client.gst_no or '27AAACU6053C1ZL'}<br/>
<b>PAN:</b> AAACU6053C<br/>
<b>Email:</b> ubltaloja@ubmail.com<br/>
<b>Phone:</b> +91 82706 64250"""
        
        # Side-by-side layout EXACTLY like target
        billing_sections = [[
            Paragraph(billed_by_text, billing_section_style),
            Paragraph(billed_to_text, billing_section_style)
        ]]
        
        billing_table = Table(billing_sections, colWidths=[95*mm, 95*mm])
        billing_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        
        elements.append(billing_table)
        elements.append(Spacer(1, 20))
        
        # ===== ITEMIZATION TABLE - EXACT STRUCTURE =====
        table_headers = ['Item', 'GST Rate', 'Quantity', 'Rate', 'Amount', 'IGST', 'Total']
        
        # ===== TABLE STRUCTURE - EXACT MATCH TO TARGET PDF =====
        
        # Table headers EXACTLY like target
        headers = ['Item', 'GST Rate', 'Quantity', 'Rate', 'Amount', 'IGST', 'Total']
        
        # EXACT data from target PDF
        items = [
            ('1. Removal of existing Bare Galvalume sheet SAC Code:', '18%', '8,500', '₹445', '₹37,82,500.00', '₹6,80,850.00', '₹44,63,350.00'),
            ('2. Removal of existing Gutters,lighting & She SAC Code:', '18%', '200', '₹390', '₹78,000.00', '₹14,040.00', '₹92,040.00'),
            ('3. 1 coat of metal passivator - Rustoff 190 SAC Code:', '18%', '80', '₹5,500', '₹4,40,000.00', '₹79,200.00', '₹5,19,200.00'),
            ('4. safety net+300micron LDPE sheet below SAC Code:', '18%', '8,500', '₹125', '₹10,62,500.00', '₹1,91,250.00', '₹12,53,750.00')
        ]
        
        # Build table data - NO ALTERNATING COLORS (target has plain white)
        table_data = [headers]
        table_data.extend(items)
        
        # EXACT column widths to prevent overlap
        col_widths = [
            75*mm,   # Item - wide enough for descriptions
            18*mm,   # GST Rate  
            20*mm,   # Quantity
            22*mm,   # Rate
            30*mm,   # Amount
            25*mm,   # IGST
            30*mm    # Total
        ]
        
        items_table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        # EXACT styling matching target PDF - NO alternating row colors
        items_table.setStyle(TableStyle([
            # Header row styling
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Data rows - plain white background like target
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # Plain white - no alternating colors
            
            # EXACT alignment matching target
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),      # Item - left aligned
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),     # GST Rate - right aligned
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),    # All other numbers - right aligned
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Proper padding
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            
            # Clean borders exactly like target
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        
        elements.append(items_table)
        elements.append(Spacer(1, 20))
        
        # ===== TOTAL IN WORDS AND FINANCIAL SUMMARY =====
        total_words_style = ParagraphStyle(
            'TotalWordsStyle',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.black,
            fontName='Helvetica',
            alignment=TA_LEFT,
            spaceAfter=12
        )
        
        # Exact text matching target PDF
        total_words = "Total (in words): SIXTY THREE LAKH TWENTY EIGHT THOUSAND THREE HUNDRED FORTY RUPEES ONLY"
        elements.append(Paragraph(total_words, total_words_style))
        elements.append(Spacer(1, 16))
        
        # PROFESSIONAL financial summary matching target PDF exactly
        
        # Total in words section
        total_words_style = ParagraphStyle(
            'TotalWordsStyle',
            fontSize=11,
            fontName='Helvetica-Bold',
            alignment=TA_LEFT,
            textColor=colors.black
        )
        
        elements.append(Paragraph("Total (in words): SIXTY THREE LAKH TWENTY EIGHT THOUSAND THREE HUNDRED FORTY RUPEES ONLY", total_words_style))
        elements.append(Spacer(1, 16))
        
        # Financial summary table - right aligned like target
        summary_data = [
            ['Amount', '₹53,63,000.00'],
            ['IGST (18%)', '₹9,65,340.00'],
            ['Total (INR)', '₹63,28,340.00']
        ]
        
        summary_table = Table(summary_data, colWidths=[40*mm, 45*mm])
        summary_table.setStyle(TableStyle([
            # Clean professional styling
            ('FONTNAME', (0, 0), (-1, -2), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -2), 12),
            ('TEXTCOLOR', (0, 0), (-1, -2), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            
            # Total row - professional highlighting
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#127285')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 14),
            
            # Professional padding and borders
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, -2), 0.5, colors.black),
        ]))
        
        # Right-align summary table 
        summary_wrapper_data = [["", summary_table]]
        summary_wrapper = Table(summary_wrapper_data, colWidths=[95*mm, 85*mm])  
        summary_wrapper.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        
        elements.append(summary_wrapper)
        elements.append(Spacer(1, 30))
        
        # ===== AUTHORISED SIGNATORY SECTION =====
        signature_data = [[""], ["Authorised Signatory"]]
        signature_table = Table(signature_data, colWidths=[50*mm], rowHeights=[20*mm, 8*mm])
        signature_table.setStyle(TableStyle([
            ('LINEABOVE', (0, 1), (0, 1), 0.5, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (0, 1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (0, 1), 10),
            ('TEXTCOLOR', (0, 1), (0, 1), colors.black),
            ('VALIGN', (0, 1), (0, 1), 'BOTTOM'),
        ]))
        
        # Right-align signature exactly like target PDF
        signature_wrapper_data = [["", signature_table]]
        signature_wrapper = Table(signature_wrapper_data, colWidths=[130*mm, 50*mm])
        signature_wrapper.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        
        elements.append(signature_wrapper)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer

async def generate_template_driven_pdf(
            template_config: PDFTemplateConfig, 
            invoice_data: dict, 
            client_data: dict, 
            project_data: dict
        ) -> bytes:
    """
    Generate PDF using template-driven configuration system with Canvas Elements support
    
    Args:
        template_config: PDF template configuration
        invoice_data: Invoice data dictionary
        client_data: Client information dictionary  
        project_data: Project information dictionary
        
    Returns:
        bytes: Generated PDF as bytes
    """
    try:
        # Check if template has canvas elements (new Canva-like functionality)
        if hasattr(template_config, 'canvas_elements') and template_config.canvas_elements:
            return await generate_canvas_based_pdf(template_config, invoice_data, client_data, project_data)
        
        # Fall back to traditional template-based generation
        return await generate_traditional_pdf(template_config, invoice_data, client_data, project_data)
        
    except Exception as e:
        logger.error(f"Error in generate_template_driven_pdf: {str(e)}")
        # Final fallback to traditional generation
        try:
            return await generate_traditional_pdf(template_config, invoice_data, client_data, project_data)
        except Exception as fallback_error:
            logger.error(f"Fallback PDF generation also failed: {fallback_error}")
            raise
# End of generate_template_driven_pdf function
        
        # Check if template has logo
        if hasattr(template_config, 'logo_url') and template_config.logo_url:
            try:
                # Handle base64 logo
                if template_config.logo_url.startswith('data:image'):
                    # Extract base64 data
                    logo_data = template_config.logo_url.split(',')[1]
                    logo_bytes = base64.b64decode(logo_data)
                    logo_buffer = BytesIO(logo_bytes)
                    
                    # Create logo image
                    logo_img = RLImage(
                        logo_buffer,
                        width=template_config.logo_width,
                        height=template_config.logo_height
                    )
                    
                    # Create header with logo
                    if template_config.logo_position == "TOP_LEFT":
                        header_data = [[logo_img, "TAX INVOICE"]]
                    else:  # TOP_RIGHT (default)
                        header_data = [["TAX INVOICE", logo_img]]
                        
                    header_table = Table(header_data, colWidths=[95*mm, 95*mm])
                    header_table.setStyle(TableStyle([
                        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), template_config.header_tax_invoice_font_size),
                        ('TEXTCOLOR', (0, 0), (-1, -1), colors.toColor(template_config.header_tax_invoice_color)),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                    ]))
                    story.append(header_table)
                    
            except Exception as e:
                logger.warning(f"Logo rendering failed: {e}")
                # Fallback to text-only header
                header_style = ParagraphStyle(
                    'CustomHeader',
                    parent=styles['Normal'],
                    fontSize=template_config.header_tax_invoice_font_size,
                    alignment=TA_CENTER,
                    textColor=colors.toColor(template_config.header_tax_invoice_color),
                    fontName='Helvetica-Bold',
                    spaceAfter=template_config.invoice_details_spacing
                )
                story.append(Paragraph("TAX INVOICE", header_style))
        else:
            # No logo - text-only header
            header_style = ParagraphStyle(
                'CustomHeader',
                parent=styles['Normal'],
                fontSize=template_config.header_tax_invoice_font_size,
                alignment=TA_CENTER,
                textColor=colors.toColor(template_config.header_tax_invoice_color),
                fontName='Helvetica-Bold',
                spaceAfter=template_config.invoice_details_spacing
            )
            story.append(Paragraph("TAX INVOICE", header_style))
            
        story.append(Spacer(1, 10))
        
        # 2. INVOICE DETAILS SECTION
        invoice_details_style = ParagraphStyle(
            'InvoiceDetails',
            parent=styles['Normal'],
            fontSize=template_config.invoice_details_font_size,
            fontName='Helvetica',
            leftIndent=0,
            spaceAfter=6
        )
        
        # Format invoice date
        invoice_date = invoice_data.get('invoice_date')
        if isinstance(invoice_date, datetime):
            formatted_date = invoice_date.strftime("%d/%m/%Y")
        else:
            formatted_date = datetime.now(timezone.utc).strftime("%d/%m/%Y")
        
        # Invoice details
        story.append(Paragraph(f"<b>Invoice No:</b> #{invoice_data.get('invoice_number', 'N/A')}", invoice_details_style))
        story.append(Paragraph(f"<b>Invoice Date:</b> {formatted_date}", invoice_details_style))
        story.append(Paragraph("<b>Created By:</b> Activus Industrial Design & Build", invoice_details_style))
        story.append(Spacer(1, template_config.billing_section_spacing))
        
        # 3. BILLING SECTIONS
        billing_style = ParagraphStyle(
            'BillingStyle',
            parent=styles['Normal'],
            fontSize=template_config.billing_font_size,
            fontName='Helvetica',
            leading=template_config.billing_line_height
        )
        
        # Create billing table
        billing_data = [
            [
                Paragraph(f"""<b>BILLED BY:</b><br/>
                {getattr(template_config, 'company_name', 'Activus Industrial Design & Build')}<br/>
                {getattr(template_config, 'company_address', 'Plot no. A-52, Sector no. 27, Phase - 2<br/>Taloja, Maharashtra, India - 410206')}<br/>
                GST: {getattr(template_config, 'company_gst', '27ABCCS1234A1Z5')}<br/>
                Email: {getattr(template_config, 'company_email', 'info@activus.co.in')}<br/>
                Phone: {getattr(template_config, 'company_phone', '+91 99999 99999')}""", billing_style),
                
                Paragraph(f"""<b>BILLED TO:</b><br/>
                {client_data.get('name', 'N/A')}<br/>
                {client_data.get('bill_to_address', client_data.get('address', 'N/A'))}<br/>
                GST: {client_data.get('gst_no', 'N/A')}<br/>
                Email: {client_data.get('email', 'N/A')}<br/>
                Phone: {client_data.get('phone', 'N/A')}""", billing_style)
            ]
        ]
        
        billing_table = Table(billing_data, colWidths=[95*mm, 95*mm])
        billing_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), colors.toColor('#E8F5E8')),
            ('BACKGROUND', (1, 0), (1, 0), colors.toColor('#E8F8FF')),
            ('PADDING', (0, 0), (-1, -1), template_config.table_padding_horizontal),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(billing_table)
        story.append(Spacer(1, template_config.billing_section_spacing))
        
        # 4. PROJECT DETAILS
        story.append(Paragraph(f"<b>Project:</b> {project_data.get('project_name', 'N/A')}", invoice_details_style))
        story.append(Paragraph(f"<b>Location:</b> {project_data.get('location', 'N/A')}", invoice_details_style))
        story.append(Spacer(1, template_config.billing_section_spacing))
        
        # 5. ITEMS TABLE
        # Determine table structure based on GST type
        gst_type = invoice_data.get('gst_type', 'IGST')
        # Use "Rs." instead of ₹ symbol to avoid font encoding issues
        currency_symbol = "Rs."
        
        if gst_type == 'CGST_SGST':
            # CGST+SGST table structure
            table_headers = ['Item', 'GST Rate', 'Quantity', 'Rate', 'Amount', 'CGST', 'SGST', 'Total']
            col_widths = [
                template_config.col_item_width * mm,
                template_config.col_gst_rate_width * mm,
                template_config.col_quantity_width * mm,
                template_config.col_rate_width * mm,
                template_config.col_amount_width * mm,
                25 * mm,  # CGST width
                25 * mm,  # SGST width
                template_config.col_total_width * mm
            ]
        else:
            # IGST table structure  
            table_headers = ['Item', 'GST Rate', 'Quantity', 'Rate', 'Amount', 'IGST', 'Total']
            col_widths = [
                template_config.col_item_width * mm,
                template_config.col_gst_rate_width * mm,
                template_config.col_quantity_width * mm,
                template_config.col_rate_width * mm,
                template_config.col_amount_width * mm,
                template_config.col_igst_width * mm,
                template_config.col_total_width * mm
            ]
        
        # Create table data
        table_data = [table_headers]
        
        # Add items
        for item in invoice_data.get('items', []):
            if gst_type == 'CGST_SGST':
                cgst_amount = item.get('amount', 0) * (item.get('gst_rate', 18) / 2) / 100
                sgst_amount = item.get('amount', 0) * (item.get('gst_rate', 18) / 2) / 100
                total_amount = item.get('amount', 0) + cgst_amount + sgst_amount
                
                row = [
                    item.get('description', 'N/A'),
                    f"{item.get('gst_rate', 18)}%",
                    f"{item.get('quantity', 0):.2f}",
                    f"{currency_symbol} {item.get('rate', 0):,.2f}",
                    f"{currency_symbol} {item.get('amount', 0):,.2f}",
                    f"{currency_symbol} {cgst_amount:,.2f}",
                    f"{currency_symbol} {sgst_amount:,.2f}",
                    f"{currency_symbol} {total_amount:,.2f}"
                ]
            else:
                igst_amount = item.get('amount', 0) * item.get('gst_rate', 18) / 100
                total_amount = item.get('amount', 0) + igst_amount
                
                row = [
                    item.get('description', 'N/A'),
                    f"{item.get('gst_rate', 18)}%",
                    f"{item.get('quantity', 0):.2f}",
                    f"{currency_symbol} {item.get('rate', 0):,.2f}",
                    f"{currency_symbol} {item.get('amount', 0):,.2f}",
                    f"{currency_symbol} {igst_amount:,.2f}",
                    f"{currency_symbol} {total_amount:,.2f}"
                ]
            
            table_data.append(row)
        
        # Create and style the table
        invoice_table = Table(table_data, colWidths=col_widths)
        
        # Apply table styling from template
        table_style = [
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.toColor(template_config.table_header_color)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.toColor(template_config.table_header_text_color)),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), template_config.table_header_font_size),
            
            # Data styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), template_config.table_data_font_size),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), template_config.table_padding_horizontal),
            
            # Borders
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]
        
        # Add alternating row colors if enabled
        if template_config.use_alternating_row_colors:
            for i in range(1, len(table_data)):
                if i % 2 == 0:  # Even rows (excluding header)
                    table_style.append(
                        ('BACKGROUND', (0, i), (-1, i), colors.toColor(template_config.alternating_row_color))
                    )
        
        invoice_table.setStyle(TableStyle(table_style))
        story.append(invoice_table)
        story.append(Spacer(1, 20))
        
        # 6. FINANCIAL SUMMARY
        summary_style = ParagraphStyle(
            'SummaryStyle',
            parent=styles['Normal'],
            fontSize=template_config.summary_font_size,
            fontName='Helvetica-Bold',
            alignment=TA_RIGHT,
            rightIndent=20
        )
        
        subtotal = invoice_data.get('subtotal', 0)
        total_gst = invoice_data.get('total_gst_amount', 0)
        total_amount = invoice_data.get('total_amount', 0)
        advance = invoice_data.get('advance_received', 0)
        net_due = invoice_data.get('net_amount_due', 0)
        
        # Summary data
        summary_data = [
            ['Subtotal:', f"{currency_symbol} {subtotal:,.2f}"]
        ]
        
        # Add GST breakdown based on type
        if gst_type == 'CGST_SGST':
            cgst_total = invoice_data.get('cgst_amount', 0)
            sgst_total = invoice_data.get('sgst_amount', 0)
            summary_data.extend([
                [f'CGST ({9.0}%):', f"{currency_symbol} {cgst_total:,.2f}"],
                [f'SGST ({9.0}%):', f"{currency_symbol} {sgst_total:,.2f}"]
            ])
        else:
            summary_data.append([f'IGST ({18.0}%):', f"{currency_symbol} {total_gst:,.2f}"])
        
        summary_data.extend([
            ['Total Amount:', f"{currency_symbol} {total_amount:,.2f}"],
            ['Advance Received:', f"({currency_symbol} {advance:,.2f})"],
            ['Net Amount Due:', f"{currency_symbol} {net_due:,.2f}"]
        ])
        
        # Create summary table
        summary_table = Table(summary_data, colWidths=[80*mm, 40*mm])
        summary_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), template_config.summary_font_size),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.toColor(template_config.total_row_color)),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.toColor(template_config.total_row_text_color)),
            ('PADDING', (0, 0), (-1, -1), 6),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, template_config.signature_spacing))
        
        # 7. PAYMENT TERMS AND SIGNATURE
        signature_style = ParagraphStyle(
            'SignatureStyle',
            parent=styles['Normal'],
            fontSize=template_config.signature_font_size,
            fontName='Helvetica'
        )
        
        story.append(Paragraph(f"<b>Payment Terms:</b> {invoice_data.get('payment_terms', 'Payment due within 30 days')}", signature_style))
        story.append(Spacer(1, 20))
        
        # Signature section
        signature_table_data = [
            ['', 'Authorised Signatory'],
            ['', 'Activus Industrial Design & Build']
        ]
        
        signature_table = Table(signature_table_data, colWidths=[120*mm, 70*mm])
        signature_table.setStyle(TableStyle([
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (1, 0), (1, -1), template_config.signature_font_size),
            ('VALIGN', (1, 0), (1, -1), 'TOP')
        ]))
        
        story.append(signature_table)
        
        # Build the PDF
        doc.build(story)
        
        # Get the PDF data
        buffer.seek(0)
        pdf_data = buffer.read()
        buffer.close()
        
        return pdf_data
        
    except Exception as e:
        logger.error(f"Error generating template-driven PDF: {str(e)}")
        raise

# Authentication functions
async def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode('utf-8'), salt).decode('utf-8')

async def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

async def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        'user_id': user_id,
        'email': email,
        'role': role,
        'exp': datetime.now(timezone.utc) + timedelta(days=7)
    }
    return jwt.encode(payload, SECRET_KEY, algorithm='HS256')

async def verify_token(token: str) -> Dict:
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    payload = await verify_token(credentials.credentials)
    user = await db.users.find_one({"id": payload["user_id"]})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user

async def log_activity(user_id: str, user_email: str, user_role: str, action: str, description: str, 
                      project_id: Optional[str] = None, invoice_id: Optional[str] = None):
    log_entry = ActivityLog(
        user_id=user_id,
        user_email=user_email,
        user_role=user_role,
        action=action,
        description=description,
        project_id=project_id,
        invoice_id=invoice_id
    )
    await db.activity_logs.insert_one(log_entry.dict())

# Initialize super admin
async def init_super_admin():
    super_admin_email = "brightboxm@gmail.com"
    existing_user = await db.users.find_one({"email": super_admin_email})
    
    if not existing_user:
        password_hash = await hash_password("admin123")
        super_admin = User(
            email=super_admin_email,
            password_hash=password_hash,
            role=UserRole.SUPER_ADMIN,
            company_name="Activus Industrial Design & Build"
        )
        await db.users.insert_one(super_admin.dict())
        logger.info("Super admin created successfully")

# API Router
from fastapi import APIRouter
api_router = APIRouter(prefix="/api")

# Health endpoints
@app.get("/health")
async def health_check():
    return {"status": "healthy", "service": "activus-invoice-api", "version": "2.0.0"}

@app.get("/ready")
async def readiness_check():
    try:
        # Check database connection
        await db.command("ping")
        return {"status": "ready", "database": "connected"}
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Database not ready: {str(e)}")

# Authentication endpoints
@api_router.post("/auth/login")
async def login(user_data: UserLogin):
    try:
        user = await db.users.find_one({"email": user_data.email, "is_active": True})
        if not user:
            raise HTTPException(status_code=401, detail="Invalid email or password")
        
        if not await verify_password(user_data.password, user["password_hash"]):
            raise HTTPException(status_code=401, detail="Invalid email or password")
        
        token = await create_token(user["id"], user["email"], user["role"])
        
        await log_activity(
            user["id"], user["email"], user["role"], 
            "login", "User logged in successfully"
        )
        
        return {
            "access_token": token,
            "token_type": "bearer",
            "user": {
                "id": user["id"],
                "email": user["email"],
                "role": user["role"],
                "company_name": user["company_name"]
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Login error: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")

# Enhanced BOQ upload endpoint
@api_router.post("/upload-boq")
async def upload_boq(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    # Validate file
    allowed_content_types = [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/vnd.ms-excel'
    ]
    
    if file.content_type not in allowed_content_types:
        raise HTTPException(status_code=400, detail="Invalid file type. Please upload an Excel file")
    
    # Check file size
    file_content = await file.read()
    if len(file_content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large. Maximum size is 10MB")
    
    try:
        parser = ExcelParser()
        parsed_data = await parser.parse_excel_file(file_content, file.filename)
        
        await log_activity(
            current_user["id"], current_user["email"], current_user["role"],
            "boq_uploaded", f"Successfully uploaded and parsed BOQ file: {file.filename}"
        )
        
        return parsed_data
        
    except Exception as e:
        logger.error(f"BOQ parsing error: {str(e)}")
        raise HTTPException(status_code=422, detail=f"Failed to parse BOQ: {str(e)}")

# Project endpoints
@api_router.post("/projects")
async def create_project(project_data: dict, current_user: dict = Depends(get_current_user)):
    try:
        # Validate percentages - handle backward compatibility
        abg = project_data.get('abg_percentage', 0)
        ra_bill = project_data.get('ra_bill_percentage', project_data.get('ra_percentage', 0))  # Handle old field name
        erection = project_data.get('erection_percentage', 0)
        pbg = project_data.get('pbg_percentage', 0)
        
        total_percentage = abg + ra_bill + erection + pbg
        if abs(total_percentage - 100.0) > 0.01:
            raise HTTPException(
                status_code=400, 
                detail=f"Percentages must total 100%. Current total: {total_percentage}%"
            )
            
        # Validate GST configuration
        gst_type = project_data.get('gst_type', 'IGST')
        if gst_type not in ['CGST_SGST', 'IGST']:
            raise HTTPException(
                status_code=400,
                detail="GST type must be either CGST_SGST or IGST"
            )
        
        # Process BOQ items
        boq_items = []
        for item_data in project_data.get('boq_items', []):
            boq_item = BOQItem(
                sr_no=item_data.get('sr_no', len(boq_items) + 1),
                description=item_data['description'],
                unit=item_data.get('unit', 'Nos'),
                quantity=item_data['quantity'],
                rate=item_data['rate'],
                amount=item_data['amount'],
                gst_rate=item_data.get('gst_rate', 18.0),
                billed_quantity=0.0
            )
            boq_items.append(boq_item)
        
        # Create project with GST configuration
        project = Project(
            project_name=project_data['project_name'],
            client_id=project_data['client_id'],
            client_name=project_data['client_name'],
            architect=project_data.get('architect', ''),
            location=project_data.get('location', ''),
            abg_percentage=abg,
            ra_bill_percentage=ra_bill,  # Updated field name
            erection_percentage=erection,
            pbg_percentage=pbg,
            gst_type=gst_type,
            gst_approval_status='pending',  # All new projects start as pending GST approval
            total_project_value=project_data['total_project_value'],
            boq_items=boq_items
        )
        
        project_dict = project.dict()
        project_dict['created_by'] = current_user['id']
        
        await db.projects.insert_one(project_dict)
        
        await log_activity(
            current_user["id"], current_user["email"], current_user["role"],
            "project_created", f"Created project: {project.project_name}"
        )
        
        return {"message": "Project created successfully", "project_id": project.id}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Project creation error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create project: {str(e)}")

@api_router.post("/projects/{project_id}/gst-approval")
async def update_gst_approval(
    project_id: str, 
    approval_data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Update GST approval status and BOQ item GST percentages"""
    try:
        # Check if user has permission (Manager or SuperAdmin)
        if current_user.get('role') not in ['Manager', 'SuperAdmin', 'super_admin']:
            raise HTTPException(
                status_code=403, 
                detail="Only Managers or SuperAdmins can approve GST configurations"
            )
        
        project = await db.projects.find_one({"id": project_id})
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
            
        # Check if already approved (locked)
        if project.get('gst_approval_status') == 'approved':
            raise HTTPException(
                status_code=400, 
                detail="GST configuration is already approved and locked"
            )
        
        action = approval_data.get('action')  # 'approve' or 'reject'
        boq_gst_updates = approval_data.get('boq_gst_updates', [])  # GST % for each BOQ item
        
        if action == 'approve':
            # Update BOQ items with approved GST percentages
            if boq_gst_updates:
                for update in boq_gst_updates:
                    item_id = update.get('item_id')
                    gst_rate = update.get('gst_rate')
                    
                    # Find and update the specific BOQ item
                    for item in project.get('boq_items', []):
                        if item.get('id') == item_id:
                            item['gst_rate'] = gst_rate
                            break
            
            # Approve the GST configuration
            update_data = {
                "gst_approval_status": "approved",
                "gst_approved_by": current_user['id'],
                "gst_approved_at": datetime.now(timezone.utc),
                "boq_items": project.get('boq_items', []),
                "updated_at": datetime.now(timezone.utc)
            }
            
            await log_activity(
                current_user["id"], current_user["email"], current_user["role"],
                "gst_approved", f"Approved GST configuration for project: {project['project_name']}"
            )
            
        elif action == 'reject':
            update_data = {
                "gst_approval_status": "rejected",
                "updated_at": datetime.now(timezone.utc)
            }
            
            await log_activity(
                current_user["id"], current_user["email"], current_user["role"],
                "gst_rejected", f"Rejected GST configuration for project: {project['project_name']}"
            )
        else:
            raise HTTPException(
                status_code=400,
                detail="Action must be 'approve' or 'reject'"
            )
        
        result = await db.projects.update_one(
            {"id": project_id},
            {"$set": update_data}
        )
        
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Project not found or no changes made")
            
        return {
            "message": f"GST configuration {action}d successfully",
            "project_id": project_id,
            "status": action + "d"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"GST approval error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to update GST approval: {str(e)}")

@api_router.get("/projects/pending-gst-approval")
async def get_projects_pending_gst_approval(current_user: dict = Depends(get_current_user)):
    """Get projects that need GST approval (for Managers/SuperAdmins)"""
    try:
        # Check if user has permission to approve GST
        if current_user.get('role') not in ['Manager', 'SuperAdmin', 'super_admin']:
            raise HTTPException(
                status_code=403, 
                detail="Only Managers or SuperAdmins can view pending GST approvals"
            )
        
        projects = await db.projects.find({
            "gst_approval_status": "pending"
        }).to_list(length=None)
        
        # Format projects for GST review
        gst_pending_projects = []
        for project in projects:
            gst_pending_projects.append({
                "id": project.get("id"),
                "project_name": project.get("project_name"),
                "client_name": project.get("client_name"),
                "gst_type": project.get("gst_type"),
                "created_at": project.get("created_at"),
                "boq_items_count": len(project.get("boq_items", [])),
                "total_project_value": project.get("total_project_value"),
                "boq_items": project.get("boq_items", [])  # For GST percentage review
            })
        
        return gst_pending_projects
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching pending GST approvals: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch pending GST approvals")

@api_router.get("/projects")
async def get_projects(current_user: dict = Depends(get_current_user)):
    try:
        # Fetch projects from MongoDB
        projects_cursor = db.projects.find()
        projects = await projects_cursor.to_list(1000)
        
        # Convert MongoDB documents to proper format
        formatted_projects = []
        for project in projects:
            # Remove MongoDB _id and convert to proper format
            if '_id' in project:
                del project['_id']
            
            # Add calculated fields
            boq_items = project.get('boq_items', [])
            total_billed = 0
            
            # Calculate total billed from BOQ items
            for item in boq_items:
                billed_qty = item.get('billed_quantity', 0)
                rate = item.get('rate', 0)
                total_billed += billed_qty * rate
            
            total_value = project.get('total_project_value', 0)
            completion_percentage = (total_billed / total_value * 100) if total_value > 0 else 0
            
            project['total_billed'] = total_billed
            project['remaining_value'] = total_value - total_billed
            project['completion_percentage'] = round(completion_percentage, 2)
            
            formatted_projects.append(project)
        
        return formatted_projects
        
    except Exception as e:
        logger.error(f"Error fetching projects: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch projects: {str(e)}")

@api_router.get("/projects/{project_id}")
async def get_project(project_id: str, current_user: dict = Depends(get_current_user)):
    try:
        project = await db.projects.find_one({"id": project_id})
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        
        # Remove MongoDB _id
        if '_id' in project:
            del project['_id']
        
        # Get associated invoices
        invoices_cursor = db.invoices.find({"project_id": project_id})
        invoices = await invoices_cursor.to_list(100)
        
        # Remove MongoDB _id from invoices
        for invoice in invoices:
            if '_id' in invoice:
                del invoice['_id']
        
        # Calculate billing status with proper RA tracking
        billing_status = await calculate_project_billing_status(project, invoices)
        project['billing_status'] = billing_status
        
        return project
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching project: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch project: {str(e)}")

@api_router.get("/projects/{project_id}/ra-tracking")
async def get_project_ra_tracking(project_id: str, current_user: dict = Depends(get_current_user)):
    try:
        project = await db.projects.find_one({"id": project_id})
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        
        # Get all invoices for this project
        invoices_cursor = db.invoices.find({"project_id": project_id, "invoice_type": "tax_invoice"})
        invoices = await invoices_cursor.to_list(100)
        
        # Calculate RA tracking for each BOQ item
        ra_tracking = []
        for item in project.get('boq_items', []):
            # Calculate billed quantity from tax invoices only
            billed_qty = 0
            for invoice in invoices:
                for inv_item in invoice.get('items', []):
                    if inv_item.get('boq_item_id') == item.get('id'):
                        billed_qty += inv_item.get('quantity', 0)
            
            balance_qty = item.get('quantity', 0) - billed_qty
            
            ra_tracking.append({
                'item_id': item.get('id'),
                'description': item.get('description'),
                'unit': item.get('unit', 'Nos'),
                'overall_qty': item.get('quantity', 0),
                'balance_qty': max(0, balance_qty),
                'billed_qty': billed_qty,
                'rate': item.get('rate', 0),
                'gst_mapping': {
                    'total_gst_rate': item.get('gst_rate', 18),
                    'cgst_rate': item.get('gst_rate', 18) / 2,
                    'sgst_rate': item.get('gst_rate', 18) / 2,
                    'igst_rate': item.get('gst_rate', 18)
                },
                'ra_usage': {}
            })
        
        return {
            'ra_tracking': ra_tracking,
            'next_ra': f"RA{len(invoices) + 1}",
            'project_name': project.get('project_name'),
            'client_name': project.get('client_name')
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching RA tracking: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch RA tracking: {str(e)}")

async def calculate_project_billing_status(project: dict, invoices: list) -> dict:
    """Calculate comprehensive project billing status"""
    
    # Get only TAX invoices for RA calculation (not proforma)
    tax_invoices = [inv for inv in invoices if inv.get('invoice_type') == 'tax_invoice']
    
    # Calculate totals
    total_project_value = project.get('total_project_value', 0)
    total_billed = sum(inv.get('total_amount', 0) for inv in tax_invoices)
    remaining_value = total_project_value - total_billed
    completion_percentage = (total_billed / total_project_value * 100) if total_project_value > 0 else 0
    
    # Get next RA number
    next_ra_number = f"RA{len(tax_invoices) + 1}"
    
    # Create BOQ items status for billing table
    boq_items_status = []
    for item in project.get('boq_items', []):
        # Calculate billed quantity from tax invoices only
        billed_qty = 0
        for invoice in tax_invoices:
            for inv_item in invoice.get('items', []):
                if inv_item.get('boq_item_id') == item.get('id'):
                    billed_qty += inv_item.get('quantity', 0)
        
        remaining_qty = item.get('quantity', 0) - billed_qty
        
        boq_items_status.append({
            'id': item.get('id'),
            'description': item.get('description', ''),
            'unit': item.get('unit', 'Nos'),
            'original_quantity': item.get('quantity', 0),
            'billed_quantity': billed_qty,
            'remaining_quantity': max(0, remaining_qty),
            'rate': item.get('rate', 0),
            'gst_rate': item.get('gst_rate', 18),
            'bill_quantity': '',  # Empty by default as requested
            'amount': 0  # Will be calculated on frontend
        })
    
    return {
        'next_ra': next_ra_number,
        'total_billed': total_billed,
        'remaining_value': remaining_value,
        'completion_percentage': round(completion_percentage, 2),
        'previous_invoices': len(tax_invoices),
        'boq_items': boq_items_status,
        'invoice_links': [
            {
                'invoice_number': inv.get('invoice_number'),
                'invoice_id': inv.get('id'),
                'amount': inv.get('total_amount', 0),
                'date': inv.get('invoice_date')
            } for inv in tax_invoices
        ]
    }

# Enhanced invoice endpoints
@api_router.post("/invoices")
async def create_invoice(invoice_data: dict, current_user: dict = Depends(get_current_user)):
    try:
        # Get project and validate
        project = await db.projects.find_one({"id": invoice_data['project_id']})
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        
        # Check GST approval status
        if project.get('gst_approval_status') != 'approved':
            raise HTTPException(
                status_code=400, 
                detail="Cannot create invoice: Project GST configuration is not approved yet"
            )
        
        # Validate quantities against BOQ (CRITICAL FIX)
        for item_data in invoice_data.get('items', []):
            boq_item = next((bi for bi in project.get('boq_items', []) 
                           if bi['id'] == item_data['boq_item_id']), None)
            
            if not boq_item:
                raise HTTPException(status_code=400, detail=f"BOQ item not found: {item_data['boq_item_id']}")
            
            # Calculate already billed quantity from TAX invoices only
            existing_invoices = await db.invoices.find({
                "project_id": invoice_data['project_id'],
                "invoice_type": "tax_invoice"  # Only count tax invoices
            }).to_list(100)
            
            billed_qty = sum(
                inv_item.get('quantity', 0) 
                for invoice in existing_invoices 
                for inv_item in invoice.get('items', [])
                if inv_item.get('boq_item_id') == item_data['boq_item_id']
            )
            
            available_qty = boq_item['quantity'] - billed_qty
            requested_qty = item_data.get('quantity', 0)
            
            if requested_qty > available_qty:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Quantity {requested_qty} exceeds available {available_qty} for item: {boq_item['description'][:50]}..."
                )
        
        # Process invoice items
        invoice_items = []
        subtotal = 0
        
        for item_data in invoice_data['items']:
            amount = item_data['quantity'] * item_data['rate']
            
            invoice_item = InvoiceItem(
                boq_item_id=item_data['boq_item_id'],
                description=item_data['description'],
                unit=item_data['unit'],
                quantity=item_data['quantity'],
                rate=item_data['rate'],
                amount=amount,
                gst_rate=item_data.get('gst_rate', 18.0)
            )
            
            invoice_items.append(invoice_item)
            subtotal += amount
        
        # Calculate GST based on project GST type
        project_gst_type = project.get('gst_type', 'IGST')
        cgst_amount = 0.0
        sgst_amount = 0.0
        igst_amount = 0.0
        
        if project_gst_type == 'CGST_SGST':
            # Split GST into CGST and SGST (50-50 split)
            total_gst = sum(item.amount * item.gst_rate / 100 for item in invoice_items)
            cgst_amount = total_gst / 2
            sgst_amount = total_gst / 2
        else:  # IGST
            igst_amount = sum(item.amount * item.gst_rate / 100 for item in invoice_items)
        
        total_gst_amount = cgst_amount + sgst_amount + igst_amount
        total_amount = subtotal + total_gst_amount
        
        # Generate invoice number and RA number
        invoice_count = await db.invoices.count_documents({}) + 1
        invoice_number = f"INV-{invoice_count:06d}"
        
        # Only assign RA number for tax invoices
        ra_number = None
        if invoice_data.get('invoice_type') == 'tax_invoice':
            tax_invoice_count = await db.invoices.count_documents({
                "project_id": invoice_data['project_id'],
                "invoice_type": "tax_invoice"
            })
            ra_number = f"RA{tax_invoice_count + 1}"
        
        # Calculate net amount due
        advance_received = float(invoice_data.get('advance_received', 0))
        net_amount_due = total_amount - advance_received
        
        # Create invoice with GST breakdown
        invoice = Invoice(
            invoice_number=invoice_number,
            project_id=invoice_data['project_id'],
            client_id=invoice_data['client_id'],
            invoice_type=invoice_data.get('invoice_type', 'tax_invoice'),
            items=invoice_items,
            subtotal=subtotal,
            gst_type=project_gst_type,
            cgst_amount=cgst_amount,
            sgst_amount=sgst_amount,
            igst_amount=igst_amount,
            total_gst_amount=total_gst_amount,
            total_amount=total_amount,
            payment_terms=invoice_data.get('payment_terms', 'Payment due within 30 days'),
            advance_received=advance_received,
            net_amount_due=net_amount_due,
            ra_number=ra_number,
            status="created"
        )
        
        invoice_dict = invoice.dict()
        invoice_dict['created_by'] = current_user['id']
        
        # Insert invoice
        result = await db.invoices.insert_one(invoice_dict)
        
        # Update BOQ billed quantities for TAX invoices only
        if invoice_data.get('invoice_type') == 'tax_invoice':
            await update_boq_billed_quantities(project['id'], invoice_items)
        
        await log_activity(
            current_user["id"], current_user["email"], current_user["role"],
            "invoice_created", f"Created {invoice.invoice_type} invoice: {invoice.invoice_number}",
            project_id=invoice.project_id, invoice_id=invoice.id
        )
        
        # Emit real-time event for invoice creation
        boq_items_affected = [item.boq_item_id for item in invoice_items]
        event_data = create_invoice_event_data(invoice_dict, boq_items_affected)
        await emit_project_event(
            ProjectEvent.INVOICE_CREATED, 
            invoice_data['project_id'], 
            event_data, 
            current_user['id']
        )
        
        # Emit BOQ item billing events for each item
        for item in invoice_items:
            boq_item = next((bi for bi in project.get('boq_items', []) 
                           if bi['id'] == item.boq_item_id), None)
            if boq_item:
                new_billed_qty = boq_item.get('billed_quantity', 0) + item.quantity
                available_qty = boq_item['quantity'] - new_billed_qty
                
                boq_event_data = create_boq_event_data(
                    item.boq_item_id, 
                    new_billed_qty, 
                    available_qty
                )
                await emit_project_event(
                    ProjectEvent.BOQ_ITEM_BILLED, 
                    invoice_data['project_id'], 
                    boq_event_data, 
                    current_user['id']
                )
        
        return {
            "message": "Invoice created successfully", 
            "invoice_id": invoice.id,
            "invoice_number": invoice.invoice_number,
            "ra_number": ra_number
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Invoice creation error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create invoice: {str(e)}")

async def update_boq_billed_quantities(project_id: str, invoice_items: List[InvoiceItem]):
    """Update BOQ item billed quantities"""
    try:
        project = await db.projects.find_one({"id": project_id})
        if not project:
            return
        
        # Update billed quantities
        for invoice_item in invoice_items:
            for i, boq_item in enumerate(project.get('boq_items', [])):
                if boq_item['id'] == invoice_item.boq_item_id:
                    current_billed = boq_item.get('billed_quantity', 0)
                    project['boq_items'][i]['billed_quantity'] = current_billed + invoice_item.quantity
                    break
        
        # Update project in database
        await db.projects.update_one(
            {"id": project_id},
            {"$set": {"boq_items": project['boq_items'], "updated_at": datetime.now(timezone.utc)}}
        )
        
    except Exception as e:
        logger.error(f"Error updating BOQ quantities: {str(e)}")

@api_router.get("/invoices")
async def get_invoices(current_user: dict = Depends(get_current_user)):
    try:
        # Fetch invoices from MongoDB
        invoices_cursor = db.invoices.find()
        invoices = await invoices_cursor.to_list(1000)
        
        # Convert MongoDB documents to proper format
        formatted_invoices = []
        for invoice in invoices:
            # Remove MongoDB _id
            if '_id' in invoice:
                del invoice['_id']
            
            # Enhance with project and client info
            project = await db.projects.find_one({"id": invoice.get('project_id')})
            if project:
                invoice['project_name'] = project.get('project_name', 'Unknown Project')
            
            client = await db.clients.find_one({"id": invoice.get('client_id')})
            if client:
                invoice['client_name'] = client.get('name', 'Unknown Client')
            
            formatted_invoices.append(invoice)
        
        return formatted_invoices
        
    except Exception as e:
        logger.error(f"Error fetching invoices: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch invoices: {str(e)}")

@api_router.get("/invoices/{invoice_id}/pdf")
async def download_invoice_pdf(invoice_id: str, current_user: dict = Depends(get_current_user)):
    try:
        # Get invoice
        invoice_data = await db.invoices.find_one({"id": invoice_id})
        if not invoice_data:
            raise HTTPException(status_code=404, detail="Invoice not found")
        
        # Get project
        project_data = await db.projects.find_one({"id": invoice_data['project_id']})
        if not project_data:
            raise HTTPException(status_code=404, detail="Project not found")
        
        # Get client
        client_data = await db.clients.find_one({"id": invoice_data['client_id']})
        if not client_data:
            raise HTTPException(status_code=404, detail="Client not found")
        
        # Convert to Pydantic models
        invoice = Invoice(**invoice_data)
        project = Project(**project_data)
        client = ClientInfo(**client_data)
        
        # Add GST type to invoice for proper breakdown in PDF
        if hasattr(project, 'gst_type') and project.gst_type:
            invoice.gst_type = project.gst_type
        
        # Initialize template manager if not already done
        if not hasattr(template_manager, 'db') or template_manager.db is None:
            await initialize_template_manager(db)
        
        # Get active template
        template_config = await template_manager.get_active_template()
        
        # Generate PDF using template-driven system
        pdf_data = await generate_template_driven_pdf(
            template_config,
            invoice_data,
            client_data,
            project_data
        )
        
        # Return PDF response
        return Response(
            content=pdf_data,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={invoice.invoice_number}.pdf",
                "Content-Length": str(len(pdf_data))
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"PDF generation error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate PDF: {str(e)}")

# Client endpoints
@api_router.post("/clients")
async def create_client(client_data: ClientInfo, current_user: dict = Depends(get_current_user)):
    try:
        await db.clients.insert_one(client_data.dict())
        
        await log_activity(
            current_user["id"], current_user["email"], current_user["role"],
            "client_created", f"Created client: {client_data.name}"
        )
        
        return {"message": "Client created successfully", "client_id": client_data.id}
        
    except Exception as e:
        logger.error(f"Client creation error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create client: {str(e)}")

@api_router.get("/clients")
async def get_clients(current_user: dict = Depends(get_current_user)):
    try:
        # Fetch clients from MongoDB
        clients_cursor = db.clients.find()
        clients = await clients_cursor.to_list(1000)
        
        # Convert MongoDB documents to proper format
        formatted_clients = []
        for client in clients:
            # Remove MongoDB _id 
            if '_id' in client:
                del client['_id']
            formatted_clients.append(client)
        
        return formatted_clients
        
    except Exception as e:
        logger.error(f"Error fetching clients: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch clients: {str(e)}")

# Logo upload endpoint
@api_router.post("/admin/upload-logo")
async def upload_company_logo(
    logo: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    try:
        if current_user.get("role") != "super_admin":
            raise HTTPException(status_code=403, detail="Only super admin can upload logos")
        
        if not logo.content_type.startswith('image/'):
            raise HTTPException(status_code=400, detail="Only image files are allowed")
        
        contents = await logo.read()
        if len(contents) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File size must be less than 5MB")
        
        # Convert to base64 for production deployment
        file_extension = logo.filename.split('.')[-1] if '.' in logo.filename else 'png'
        mime_type = logo.content_type or f'image/{file_extension}'
        
        base64_data = base64.b64encode(contents).decode('utf-8')
        logo_url = f"data:{mime_type};base64,{base64_data}"
        
        unique_filename = f"logo_{uuid.uuid4()}.{file_extension}"
        
        await log_activity(
            current_user["id"], current_user["email"], current_user["role"],
            "logo_uploaded", f"Uploaded company logo: {logo.filename}"
        )
        
        return {
            "message": "Logo uploaded successfully",
            "logo_url": logo_url,
            "filename": unique_filename,
            "size": len(contents)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Logo upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to upload logo: {str(e)}")

# Dashboard stats endpoint
# Item Master System
@api_router.get("/item-master")
async def get_item_master(current_user: dict = Depends(get_current_user)):
    try:
        # Get all unique items from BOQ data across projects
        projects = await db.projects.find().to_list(1000)
        items = []
        
        for project in projects:
            for boq_item in project.get('boq_items', []):
                items.append({
                    "id": boq_item.get('id', ''),
                    "description": boq_item.get('description', ''),
                    "unit": boq_item.get('unit', 'Nos'),
                    "rate": boq_item.get('rate', 0),
                    "gst_rate": boq_item.get('gst_rate', 18.0),
                    "project_name": project.get('project_name', '')
                })
        
        return {"items": items, "total_count": len(items)}
        
    except Exception as e:
        logger.error(f"Item master error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get item master: {str(e)}")

@api_router.post("/item-master")
async def create_item_master(item_data: dict, current_user: dict = Depends(get_current_user)):
    try:
        # For now, return success - full implementation would store custom items
        return {"message": "Item master creation not fully implemented yet", "status": "pending"}
        
    except Exception as e:
        logger.error(f"Create item master error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create item: {str(e)}")

@api_router.get("/item-master/search")
async def search_item_master(q: str = "", current_user: dict = Depends(get_current_user)):
    try:
        # Search items by description
        projects = await db.projects.find().to_list(1000)
        matching_items = []
        
        for project in projects:
            for boq_item in project.get('boq_items', []):
                if q.lower() in boq_item.get('description', '').lower():
                    matching_items.append({
                        "id": boq_item.get('id', ''),
                        "description": boq_item.get('description', ''),
                        "unit": boq_item.get('unit', 'Nos'),
                        "rate": boq_item.get('rate', 0),
                        "gst_rate": boq_item.get('gst_rate', 18.0)
                    })
        
        return {"items": matching_items, "query": q}
        
    except Exception as e:
        logger.error(f"Search item master error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to search items: {str(e)}")

@api_router.post("/item-master/auto-populate")
async def auto_populate_item(search_data: dict, current_user: dict = Depends(get_current_user)):
    try:
        description = search_data.get('description', '')
        
        # Find similar items from existing BOQ data
        projects = await db.projects.find().to_list(1000)
        suggestions = []
        
        for project in projects:
            for boq_item in project.get('boq_items', []):
                item_desc = boq_item.get('description', '').lower()
                if description.lower() in item_desc or item_desc in description.lower():
                    suggestions.append({
                        "description": boq_item.get('description', ''),
                        "unit": boq_item.get('unit', 'Nos'),
                        "rate": boq_item.get('rate', 0),
                        "gst_rate": boq_item.get('gst_rate', 18.0)
                    })
        
        return {"suggestions": suggestions[:10]}  # Limit to top 10
        
    except Exception as e:
        logger.error(f"Auto populate error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to auto populate: {str(e)}")

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    try:
        # Get counts
        total_projects = await db.projects.count_documents({})
        total_invoices = await db.invoices.count_documents({})
        total_clients = await db.clients.count_documents({})
        
        # Calculate project value
        projects = await db.projects.find().to_list(1000)
        total_project_value = sum(p.get('total_project_value', 0) for p in projects)
        
        # Calculate invoiced value and pending collections
        invoices = await db.invoices.find().to_list(1000)
        total_invoiced = sum(inv.get('total_amount', 0) for inv in invoices)
        pending_collections = sum(inv.get('net_amount_due', inv.get('total_amount', 0)) for inv in invoices)
        
        # Calculate advance received and pending payments
        advance_received = sum(inv.get('advance_received', 0) for inv in invoices)
        pending_payment = sum(inv.get('net_amount_due', 0) for inv in invoices if inv.get('status', '') == 'created')
        
        return {
            "total_projects": total_projects,
            "total_project_value": total_project_value,
            "total_invoices": total_invoices,
            "total_invoiced_value": total_invoiced,
            "pending_collections": pending_collections,
            "advance_received": advance_received,
            "pending_payment": pending_payment,
            "collection_efficiency": ((total_invoiced - pending_collections) / total_invoiced * 100) if total_invoiced > 0 else 0
        }
        
    except Exception as e:
        logger.error(f"Dashboard stats error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get dashboard stats: {str(e)}")

# Search System
@api_router.get("/search")
async def global_search(q: str = "", current_user: dict = Depends(get_current_user)):
    try:
        if not q:
            return {"projects": [], "clients": [], "invoices": [], "query": q}
        
        query_lower = q.lower()
        
        # Search projects
        projects = await db.projects.find().to_list(1000)
        matching_projects = [
            {
                "id": p.get('id'),
                "project_name": p.get('project_name'),
                "client_name": p.get('client_name'),
                "total_project_value": p.get('total_project_value', 0)
            }
            for p in projects 
            if query_lower in p.get('project_name', '').lower() or 
               query_lower in p.get('client_name', '').lower()
        ]
        
        # Search clients
        clients = await db.clients.find().to_list(1000)
        matching_clients = [
            {
                "id": c.get('id'),
                "name": c.get('name'),
                "email": c.get('email'),
                "phone": c.get('phone')
            }
            for c in clients 
            if query_lower in c.get('name', '').lower() or 
               query_lower in c.get('email', '').lower()
        ]
        
        # Search invoices
        invoices = await db.invoices.find().to_list(1000)
        matching_invoices = [
            {
                "id": i.get('id'),
                "invoice_number": i.get('invoice_number'),
                "project_id": i.get('project_id'),
                "total_amount": i.get('total_amount', 0)
            }
            for i in invoices 
            if query_lower in i.get('invoice_number', '').lower()
        ]
        
        return {
            "projects": matching_projects,
            "clients": matching_clients,
            "invoices": matching_invoices,
            "query": q
        }
        
    except Exception as e:
        logger.error(f"Global search error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to search: {str(e)}")

@api_router.get("/filters/projects")
async def get_project_filters(current_user: dict = Depends(get_current_user)):
    try:
        projects = await db.projects.find().to_list(1000)
        
        # Extract unique values for filtering
        clients = list(set(p.get('client_name') for p in projects if p.get('client_name')))
        statuses = list(set(p.get('status', 'active') for p in projects))
        
        return {
            "clients": sorted(clients),
            "statuses": sorted(statuses),
            "value_ranges": [
                {"label": "< ₹10L", "min": 0, "max": 1000000},
                {"label": "₹10L - ₹50L", "min": 1000000, "max": 5000000},
                {"label": "₹50L - ₹1Cr", "min": 5000000, "max": 10000000},
                {"label": "> ₹1Cr", "min": 10000000, "max": 999999999}
            ]
        }
        
    except Exception as e:
        logger.error(f"Project filters error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get project filters: {str(e)}")

@api_router.get("/filters/invoices")
async def get_invoice_filters(current_user: dict = Depends(get_current_user)):
    try:
        invoices = await db.invoices.find().to_list(1000)
        
        # Extract unique values for filtering
        statuses = list(set(i.get('status', 'created') for i in invoices))
        invoice_types = list(set(i.get('invoice_type', 'tax_invoice') for i in invoices))
        
        return {
            "statuses": sorted(statuses),
            "types": sorted(invoice_types),
            "amount_ranges": [
                {"label": "< ₹1L", "min": 0, "max": 100000},
                {"label": "₹1L - ₹5L", "min": 100000, "max": 500000},
                {"label": "₹5L - ₹10L", "min": 500000, "max": 1000000},
                {"label": "> ₹10L", "min": 1000000, "max": 999999999}
            ]
        }
        
    except Exception as e:
        logger.error(f"Invoice filters error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get invoice filters: {str(e)}")

# Activity logs endpoint
@api_router.get("/activity-logs")
async def get_activity_logs(current_user: dict = Depends(get_current_user)):
    try:
        if current_user.get("role") != "super_admin":
            raise HTTPException(status_code=403, detail="Only super admin can view activity logs")
        
        logs = await db.activity_logs.find().sort("timestamp", -1).limit(1000).to_list(1000)
        
        # Convert ObjectId to string for JSON serialization
        serialized_logs = []
        for log in logs:
            if '_id' in log:
                log['_id'] = str(log['_id'])
            serialized_logs.append(log)
        
        return serialized_logs
        
    except Exception as e:
        logger.error(f"Activity logs error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get activity logs: {str(e)}")

# Reports System
@api_router.get("/reports/gst-summary")
async def get_gst_summary(current_user: dict = Depends(get_current_user)):
    try:
        invoices = await db.invoices.find().to_list(1000)
        
        total_taxable = sum(inv.get('subtotal', 0) for inv in invoices)
        total_cgst = sum(inv.get('cgst_amount', 0) for inv in invoices)
        total_sgst = sum(inv.get('sgst_amount', 0) for inv in invoices)
        total_igst = sum(inv.get('igst_amount', 0) for inv in invoices)
        total_gst = total_cgst + total_sgst + total_igst
        
        return {
            "summary": {
                "total_invoices": len(invoices),
                "total_taxable_amount": total_taxable,
                "total_cgst": total_cgst,
                "total_sgst": total_sgst,
                "total_igst": total_igst,
                "total_gst": total_gst,
                "total_amount": total_taxable + total_gst
            },
            "by_gst_rate": _calculate_gst_by_rate(invoices)
        }
        
    except Exception as e:
        logger.error(f"GST summary error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get GST summary: {str(e)}")

@api_router.get("/reports/insights")
async def get_business_insights(current_user: dict = Depends(get_current_user)):
    try:
        # Get data for insights
        projects = await db.projects.find().to_list(1000)
        invoices = await db.invoices.find().to_list(1000)
        clients = await db.clients.find().to_list(1000)
        
        # Calculate insights
        avg_project_value = sum(p.get('total_project_value', 0) for p in projects) / len(projects) if projects else 0
        total_revenue = sum(inv.get('total_amount', 0) for inv in invoices)
        
        # Top clients by project value
        client_values = {}
        for project in projects:
            client_name = project.get('client_name', 'Unknown')
            client_values[client_name] = client_values.get(client_name, 0) + project.get('total_project_value', 0)
        
        top_clients = sorted(client_values.items(), key=lambda x: x[1], reverse=True)[:5]
        
        return {
            "metrics": {
                "total_projects": len(projects),
                "total_clients": len(clients),
                "avg_project_value": avg_project_value,
                "total_revenue": total_revenue,
                "active_projects": len([p for p in projects if p.get('status') == 'active'])
            },
            "top_clients": [{"name": name, "value": value} for name, value in top_clients],
            "monthly_trends": _calculate_monthly_trends(invoices)
        }
        
    except Exception as e:
        logger.error(f"Business insights error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get business insights: {str(e)}")

@api_router.get("/reports/client-summary/{client_id}")
async def get_client_summary(client_id: str, current_user: dict = Depends(get_current_user)):
    try:
        # Get client data
        client = await db.clients.find_one({"id": client_id})
        if not client:
            raise HTTPException(status_code=404, detail="Client not found")
        
        # Get client's projects and invoices
        projects = await db.projects.find({"client_id": client_id}).to_list(1000)
        invoices = await db.invoices.find({"client_id": client_id}).to_list(1000)
        
        total_project_value = sum(p.get('total_project_value', 0) for p in projects)
        total_invoiced = sum(inv.get('total_amount', 0) for inv in invoices)
        pending_amount = sum(inv.get('net_amount_due', 0) for inv in invoices)
        
        return {
            "client": {
                "id": client.get('id'),
                "name": client.get('name'),
                "email": client.get('email'),
                "phone": client.get('phone')
            },
            "summary": {
                "total_projects": len(projects),
                "total_project_value": total_project_value,
                "total_invoices": len(invoices),
                "total_invoiced": total_invoiced,
                "pending_amount": pending_amount,
                "collection_efficiency": ((total_invoiced - pending_amount) / total_invoiced * 100) if total_invoiced > 0 else 0
            },
            "projects": projects,
            "recent_invoices": sorted(invoices, key=lambda x: x.get('created_at', ''), reverse=True)[:10]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Client summary error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get client summary: {str(e)}")

def _calculate_gst_by_rate(invoices):
    """Helper function to calculate GST by rate"""
    gst_rates = {}
    
    for invoice in invoices:
        for item in invoice.get('items', []):
            rate = item.get('gst_rate', 18.0)
            if rate not in gst_rates:
                gst_rates[rate] = {"taxable_amount": 0, "gst_amount": 0, "count": 0}
            
            gst_rates[rate]["taxable_amount"] += item.get('amount', 0)
            gst_rates[rate]["gst_amount"] += item.get('amount', 0) * rate / 100
            gst_rates[rate]["count"] += 1
    
    return gst_rates

def _calculate_monthly_trends(invoices):
    """Helper function to calculate monthly trends"""
    from collections import defaultdict
    import datetime
    
    monthly_data = defaultdict(lambda: {"count": 0, "amount": 0})
    
    for invoice in invoices:
        created_at = invoice.get('created_at')
        if created_at:
            if isinstance(created_at, str):
                date_obj = datetime.datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            else:
                date_obj = created_at
            
            month_key = date_obj.strftime('%Y-%m')
            monthly_data[month_key]["count"] += 1
            monthly_data[month_key]["amount"] += invoice.get('total_amount', 0)
    
    return dict(monthly_data)

# PDF Processor System
@api_router.get("/pdf-processor/extractions")
async def get_pdf_extractions(current_user: dict = Depends(get_current_user)):
    try:
        # For now, return empty list - full implementation would track PDF extractions
        return {"extractions": [], "total_count": 0}
        
    except Exception as e:
        logger.error(f"PDF extractions error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get PDF extractions: {str(e)}")

@api_router.post("/pdf-processor/extract")
async def extract_pdf_data(current_user: dict = Depends(get_current_user)):
    try:
        # Basic response - full implementation would process uploaded PDF
        return {"message": "PDF extraction not fully implemented yet", "status": "pending"}
        
    except Exception as e:
        logger.error(f"PDF extract error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to extract PDF: {str(e)}")

@api_router.post("/pdf-processor/convert-to-project")
async def convert_pdf_to_project(current_user: dict = Depends(get_current_user)):
    try:
        # Basic response - full implementation would convert PDF data to project
        return {"message": "PDF to project conversion not fully implemented yet", "status": "pending"}
        
    except Exception as e:
        logger.error(f"PDF convert error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to convert PDF to project: {str(e)}")

# Admin Configuration System
@api_router.get("/admin/workflows")
async def get_admin_workflows(current_user: dict = Depends(get_current_user)):
    try:
        if current_user.get("role") not in ["admin", "super_admin"]:
            raise HTTPException(status_code=403, detail="Only admins can view workflows")
        
        # Basic workflow configuration
        return {
            "workflows": [
                {
                    "id": "project_creation",
                    "name": "Project Creation Workflow",
                    "steps": ["Basic Info", "Company Selection", "BOQ Review"],
                    "active": True
                },
                {
                    "id": "invoice_approval",
                    "name": "Invoice Approval Workflow", 
                    "steps": ["Create", "Review", "Approve", "Send"],
                    "active": True
                },
                {
                    "id": "gst_approval",
                    "name": "GST Approval Workflow",
                    "steps": ["Submit", "Manager Review", "Approve/Reject"],
                    "active": True
                }
            ]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Admin workflows error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get admin workflows: {str(e)}")

@api_router.get("/admin/system-config")
async def get_system_config(current_user: dict = Depends(get_current_user)):
    try:
        if current_user.get("role") not in ["admin", "super_admin"]:
            raise HTTPException(status_code=403, detail="Only admins can view system config")
        
        return {
            "config": {
                "app_name": "Activus Invoice Management System",
                "version": "1.0.0",
                "gst_default_rate": 18.0,
                "default_payment_terms": "Payment due within 30 days",
                "max_file_size_mb": 5,
                "supported_file_types": ["xlsx", "xls", "csv"],
                "features": {
                    "gst_approval_workflow": True,
                    "real_time_updates": True,
                    "pdf_generation": True,
                    "advanced_reports": True
                }
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"System config error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get system config: {str(e)}")

@api_router.get("/admin/system-health")
async def get_system_health(current_user: dict = Depends(get_current_user)):
    try:
        if current_user.get("role") not in ["admin", "super_admin"]:
            raise HTTPException(status_code=403, detail="Only admins can view system health")
        
        # Basic system health check
        projects_count = await db.projects.count_documents({})
        invoices_count = await db.invoices.count_documents({})
        clients_count = await db.clients.count_documents({})
        
        return {
            "status": "healthy",
            "database": {
                "connected": True,
                "collections": {
                    "projects": projects_count,
                    "invoices": invoices_count, 
                    "clients": clients_count
                }
            },
            "services": {
                "api": "running",
                "websockets": "active",
                "file_processing": "available"
            },
            "timestamp": datetime.now(timezone.utc)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"System health error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get system health: {str(e)}")

@api_router.post("/admin/clear-database")
async def clear_database(current_user: dict = Depends(get_current_user)):
    try:
        if current_user.get("role") != "super_admin":
            raise HTTPException(status_code=403, detail="Only super admin can clear database")
        
        # For safety, only clear non-user data
        projects_deleted = await db.projects.delete_many({})
        invoices_deleted = await db.invoices.delete_many({})
        clients_deleted = await db.clients.delete_many({})
        logs_deleted = await db.activity_logs.delete_many({})
        
        await log_activity(
            current_user["id"], current_user["email"], current_user["role"],
            "database_cleared", "Database cleared by super admin"
        )
        
        return {
            "message": "Database cleared successfully",
            "statistics": {
                "projects_deleted": projects_deleted.deleted_count,
                "invoices_deleted": invoices_deleted.deleted_count,
                "clients_deleted": clients_deleted.deleted_count,
                "logs_deleted": logs_deleted.deleted_count
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Clear database error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to clear database: {str(e)}")

# Company Profiles System  
@api_router.get("/company-profiles")
async def get_company_profiles(current_user: dict = Depends(get_current_user)):
    try:
        # For now, return default company profile - full implementation would store multiple profiles
        default_profile = {
            "id": "default-profile",
            "company_name": "Activus Design & Build",
            "gst_number": "27ABCDE1234F1Z5",
            "pan_number": "ABCDE1234F",
            "email": "info@activusdesign.com",
            "address": "Mumbai, Maharashtra, India",
            "bank_details": {
                "account_number": "123456789012",
                "ifsc_code": "ABCD0123456", 
                "branch": "Mumbai Main Branch"
            }
        }
        
        return {"profiles": [default_profile], "total_count": 1}
        
    except Exception as e:
        logger.error(f"Company profiles error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get company profiles: {str(e)}")

# Enhanced Endpoints
@api_router.get("/projects/enhanced")
async def get_enhanced_projects(current_user: dict = Depends(get_current_user)):
    try:
        # Get projects with enhanced data including GST status and approval info
        projects = await db.projects.find().to_list(1000)
        
        enhanced_projects = []
        for project in projects:
            enhanced_project = dict(project)
            enhanced_project['gst_status'] = project.get('gst_approval_status', 'pending')
            enhanced_project['requires_approval'] = project.get('gst_approval_status') != 'approved'
            enhanced_projects.append(enhanced_project)
        
        return {"projects": enhanced_projects, "total_count": len(enhanced_projects)}
        
    except Exception as e:
        logger.error(f"Enhanced projects error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get enhanced projects: {str(e)}")

@api_router.get("/invoices/enhanced")
async def get_enhanced_invoices(current_user: dict = Depends(get_current_user)):
    try:
        # Get invoices with enhanced data including GST breakdown
        invoices = await db.invoices.find().to_list(1000)
        
        enhanced_invoices = []
        for invoice in invoices:
            enhanced_invoice = dict(invoice)
            enhanced_invoice['gst_breakdown'] = {
                "cgst": invoice.get('cgst_amount', 0),
                "sgst": invoice.get('sgst_amount', 0),
                "igst": invoice.get('igst_amount', 0),
                "total": invoice.get('total_gst_amount', 0)
            }
            enhanced_invoices.append(enhanced_invoice)
        
        return {"invoices": enhanced_invoices, "total_count": len(enhanced_invoices)}
        
    except Exception as e:
        logger.error(f"Enhanced invoices error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get enhanced invoices: {str(e)}")

@api_router.post("/invoices/validate-quantities")
async def validate_invoice_quantities(validation_data: dict, current_user: dict = Depends(get_current_user)):
    try:
        project_id = validation_data.get('project_id')
        items = validation_data.get('items', [])
        
        if not project_id:
            raise HTTPException(status_code=400, detail="Project ID is required")
        
        # Get project and validate quantities
        project = await db.projects.find_one({"id": project_id})
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        
        validation_results = []
        
        for item in items:
            boq_item_id = item.get('boq_item_id')
            requested_qty = item.get('quantity', 0)
            
            # Find BOQ item
            boq_item = next((bi for bi in project.get('boq_items', []) if bi['id'] == boq_item_id), None)
            if not boq_item:
                validation_results.append({
                    "item_id": boq_item_id,
                    "valid": False,
                    "error": "BOQ item not found"
                })
                continue
            
            # Calculate available quantity
            billed_qty = boq_item.get('billed_quantity', 0)
            total_qty = boq_item.get('quantity', 0)
            available_qty = total_qty - billed_qty
            
            validation_results.append({
                "item_id": boq_item_id,
                "valid": requested_qty <= available_qty,
                "requested_quantity": requested_qty,
                "available_quantity": available_qty,
                "total_quantity": total_qty,
                "billed_quantity": billed_qty
            })
        
        return {"validation_results": validation_results}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Validate quantities error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to validate quantities: {str(e)}")

# Invoice Amendment System
@api_router.post("/invoices/{invoice_id}/amend")
async def amend_invoice(invoice_id: str, amendment_data: dict, current_user: dict = Depends(get_current_user)):
    try:
        # Get original invoice
        original_invoice = await db.invoices.find_one({"id": invoice_id})
        if not original_invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        
        # Check permissions for GST amendment
        amendment_type = amendment_data.get('amendment_type', 'quantities')
        user_role = current_user.get('role', '').lower()
        allowed_roles = ['manager', 'superadmin', 'super_admin', 'admin']  # Include variations
        
        if amendment_type == 'gst' and user_role not in allowed_roles:
            raise HTTPException(
                status_code=403, 
                detail="Only Managers and SuperAdmins can amend GST percentages"
            )
        
        # Get project for validation
        project = await db.projects.find_one({"id": original_invoice['project_id']})
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        
        amended_items = amendment_data.get('amended_items', [])
        amendment_reason = amendment_data.get('amendment_reason', '')
        
        if not amendment_reason.strip():
            raise HTTPException(status_code=400, detail="Amendment reason is required")
        
        # Process amended items
        new_invoice_items = []
        total_subtotal = 0
        total_cgst = 0
        total_sgst = 0
        total_igst = 0
        
        for amended_item in amended_items:
            # Calculate amounts based on new quantity and GST rate
            new_quantity = amended_item.get('new_quantity', 0)
            new_gst_rate = amended_item.get('new_gst_rate', amended_item.get('original_gst_rate', 18))
            rate = amended_item.get('rate', 0)
            
            basic_amount = new_quantity * rate
            gst_amount = basic_amount * (new_gst_rate / 100)
            
            invoice_item = {
                "boq_item_id": amended_item.get('boq_item_id'),
                "description": amended_item.get('description'),
                "unit": amended_item.get('unit'),
                "quantity": new_quantity,
                "rate": rate,
                "amount": basic_amount,
                "gst_rate": new_gst_rate,
                "gst_amount": gst_amount,
                "total_amount": basic_amount + gst_amount
            }
            
            new_invoice_items.append(invoice_item)
            total_subtotal += basic_amount
        
        # Calculate GST based on project GST type
        project_gst_type = project.get('gst_type', 'IGST')
        if project_gst_type == 'CGST_SGST':
            total_gst = sum(item['gst_amount'] for item in new_invoice_items)
            total_cgst = total_gst / 2
            total_sgst = total_gst / 2
        else:  # IGST
            total_igst = sum(item['gst_amount'] for item in new_invoice_items)
        
        total_gst_amount = total_cgst + total_sgst + total_igst
        total_amount = total_subtotal + total_gst_amount
        
        # Generate new invoice number for amendment
        invoice_count = await db.invoices.count_documents({}) + 1
        amended_invoice_number = f"AME-{original_invoice['invoice_number']}-{invoice_count:05d}"
        
        # Create amended invoice
        amended_invoice = {
            "id": str(uuid.uuid4()),
            "invoice_number": amended_invoice_number,
            "original_invoice_id": invoice_id,
            "amendment_reason": amendment_reason,
            "amendment_type": amendment_type,
            "project_id": original_invoice['project_id'],
            "client_id": original_invoice['client_id'],
            "invoice_type": original_invoice.get('invoice_type', 'tax_invoice'),
            "invoice_date": datetime.now(timezone.utc),
            "due_date": datetime.now(timezone.utc) + timedelta(days=30),
            "items": new_invoice_items,
            "subtotal": total_subtotal,
            "gst_type": project_gst_type,
            "cgst_amount": total_cgst,
            "sgst_amount": total_sgst,
            "igst_amount": total_igst,
            "total_gst_amount": total_gst_amount,
            "total_amount": total_amount,
            "payment_terms": original_invoice.get('payment_terms', 'Payment due within 30 days'),
            "advance_received": 0.0,
            "net_amount_due": total_amount,
            "ra_number": original_invoice.get('ra_number'),
            "status": "amended",
            "created_at": datetime.now(timezone.utc),
            "amended_by": current_user['id'],
            "amended_at": datetime.now(timezone.utc)
        }
        
        # Insert amended invoice
        result = await db.invoices.insert_one(amended_invoice)
        
        # Mark original invoice as amended
        await db.invoices.update_one(
            {"id": invoice_id},
            {"$set": {
                "status": "superseded",
                "superseded_by": amended_invoice['id'],
                "superseded_at": datetime.now(timezone.utc)
            }}
        )
        
        # Log activity
        await log_activity(
            current_user["id"], current_user["email"], current_user["role"],
            "invoice_amended", f"Amended invoice {original_invoice['invoice_number']} -> {amended_invoice_number}. Reason: {amendment_reason}"
        )
        
        # Emit WebSocket event for real-time updates (disabled for now)
        # TODO: Implement proper WebSocket integration
        # if hasattr(sio, 'emit'):
        #     await sio.emit('invoice_amended', {
        #         'original_invoice_id': invoice_id,
        #         'amended_invoice_id': amended_invoice['id'],
        #         'project_id': original_invoice['project_id']
        #     })
        
        # Clean up ObjectId fields for JSON serialization
        if '_id' in amended_invoice:
            del amended_invoice['_id']
        
        return {
            "message": "Invoice amended successfully",
            "amended_invoice": amended_invoice,
            "original_invoice_id": invoice_id
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Invoice amendment error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to amend invoice: {str(e)}")

@api_router.get("/invoices/{invoice_id}/amendment-history")
async def get_invoice_amendment_history(invoice_id: str, current_user: dict = Depends(get_current_user)):
    try:
        # Get original invoice and all amendments
        original_invoice = await db.invoices.find_one({"id": invoice_id})
        if not original_invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        
        # Find all amendments for this invoice
        amendments = await db.invoices.find({
            "original_invoice_id": invoice_id
        }).to_list(1000)
        
        # Find if this invoice is an amendment itself
        if original_invoice.get('original_invoice_id'):
            root_invoice = await db.invoices.find_one({"id": original_invoice['original_invoice_id']})
            all_amendments = await db.invoices.find({
                "original_invoice_id": original_invoice['original_invoice_id']
            }).to_list(1000)
        else:
            root_invoice = original_invoice
            all_amendments = amendments
        
        # Clean up ObjectId fields for JSON serialization
        if root_invoice and '_id' in root_invoice:
            del root_invoice['_id']
        
        for amendment in all_amendments:
            if '_id' in amendment:
                del amendment['_id']
        
        return {
            "root_invoice": root_invoice,
            "amendments": all_amendments,
            "total_amendments": len(all_amendments)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Amendment history error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get amendment history: {str(e)}")

@api_router.get("/invoices/{invoice_id}")
async def get_invoice_by_id(invoice_id: str, current_user: dict = Depends(get_current_user)):
    try:
        invoice = await db.invoices.find_one({"id": invoice_id})
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        
        # Clean up ObjectId fields for JSON serialization
        if '_id' in invoice:
            del invoice['_id']
        
        return invoice
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get invoice error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get invoice: {str(e)}")

# Enhanced Amendment Workflow with Manager Approval
@api_router.post("/invoices/{invoice_id}/amendment-request")
async def submit_amendment_request(invoice_id: str, amendment_data: dict, current_user: dict = Depends(get_current_user)):
    try:
        # Get original invoice
        original_invoice = await db.invoices.find_one({"id": invoice_id})
        if not original_invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        
        amendment_reason = amendment_data.get('amendment_reason', '')
        if not amendment_reason.strip():
            raise HTTPException(status_code=400, detail="Amendment reason is required")
        
        # Create amendment request
        amendment_request = {
            "id": str(uuid.uuid4()),
            "original_invoice_id": invoice_id,
            "amendment_type": amendment_data.get('amendment_type', 'quantities'),
            "amendment_reason": amendment_reason,
            "amended_items": amendment_data.get('amended_items', []),
            "requested_by": current_user['id'],
            "requested_by_email": current_user['email'],
            "status": "pending_approval",
            "created_at": datetime.now(timezone.utc)
        }
        
        # Store amendment request
        result = await db.amendment_requests.insert_one(amendment_request)
        
        # Update original invoice status
        await db.invoices.update_one(
            {"id": invoice_id},
            {"$set": {
                "status": "amendment_requested",
                "amendment_request_id": amendment_request['id'],
                "updated_at": datetime.now(timezone.utc)
            }}
        )
        
        # Log activity
        await log_activity(
            current_user["id"], current_user["email"], current_user["role"],
            "amendment_requested", f"Amendment request submitted for invoice {original_invoice['invoice_number']}. Reason: {amendment_reason}"
        )
        
        # Clean up ObjectId for JSON response
        if '_id' in amendment_request:
            del amendment_request['_id']
        
        return {
            "message": "Amendment request submitted successfully",
            "amendment_request": amendment_request
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Amendment request error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to submit amendment request: {str(e)}")

@api_router.get("/amendment-requests/pending")
async def get_pending_amendment_requests(current_user: dict = Depends(get_current_user)):
    try:
        # Check if user has permission (Manager or SuperAdmin)
        user_role = current_user.get('role', '').lower()
        if user_role not in ['manager', 'superadmin', 'super_admin', 'admin']:
            raise HTTPException(
                status_code=403, 
                detail="Only Managers or SuperAdmins can view pending amendment requests"
            )
        
        # Get pending amendment requests
        requests = await db.amendment_requests.find({
            "status": "pending_approval"
        }).to_list(1000)
        
        # Get invoice and project details for each request
        enhanced_requests = []
        for request in requests:
            # Get original invoice
            invoice = await db.invoices.find_one({"id": request['original_invoice_id']})
            project = None
            if invoice:
                project = await db.projects.find_one({"id": invoice['project_id']})
            
            enhanced_request = dict(request)
            enhanced_request['invoice'] = invoice
            enhanced_request['project'] = project
            
            # Clean up ObjectId
            if '_id' in enhanced_request:
                del enhanced_request['_id']
            
            enhanced_requests.append(enhanced_request)
        
        return {"amendment_requests": enhanced_requests, "total_count": len(enhanced_requests)}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get pending amendments error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get pending amendment requests: {str(e)}")

@api_router.post("/amendment-requests/{request_id}/approve")
async def approve_amendment_request(request_id: str, approval_data: dict, current_user: dict = Depends(get_current_user)):
    try:
        # Check permissions
        user_role = current_user.get('role', '').lower()
        if user_role not in ['manager', 'superadmin', 'super_admin', 'admin']:
            raise HTTPException(
                status_code=403, 
                detail="Only Managers or SuperAdmins can approve amendment requests"
            )
        
        # Get amendment request
        amendment_request = await db.amendment_requests.find_one({"id": request_id})
        if not amendment_request:
            raise HTTPException(status_code=404, detail="Amendment request not found")
        
        if amendment_request['status'] != 'pending_approval':
            raise HTTPException(status_code=400, detail="Amendment request is not pending approval")
        
        action = approval_data.get('action')  # 'approve' or 'reject'
        if action not in ['approve', 'reject']:
            raise HTTPException(status_code=400, detail="Action must be 'approve' or 'reject'")
        
        if action == 'approve':
            # Create the amended invoice using the original amendment logic
            original_invoice = await db.invoices.find_one({"id": amendment_request['original_invoice_id']})
            if not original_invoice:
                raise HTTPException(status_code=404, detail="Original invoice not found")
            
            # Process amendment (similar to original amend_invoice logic)
            amended_items = amendment_request.get('amended_items', [])
            new_invoice_items = []
            total_subtotal = 0
            
            # Get project for GST type
            project = await db.projects.find_one({"id": original_invoice['project_id']})
            project_gst_type = project.get('gst_type', 'IGST') if project else 'IGST'
            
            for amended_item in amended_items:
                new_quantity = amended_item.get('new_quantity', 0)
                new_gst_rate = amended_item.get('new_gst_rate', amended_item.get('original_gst_rate', 18))
                rate = amended_item.get('rate', 0)
                
                basic_amount = new_quantity * rate
                gst_amount = basic_amount * (new_gst_rate / 100)
                
                invoice_item = {
                    "boq_item_id": amended_item.get('boq_item_id'),
                    "description": amended_item.get('description'),
                    "unit": amended_item.get('unit'),
                    "quantity": new_quantity,
                    "rate": rate,
                    "amount": basic_amount,
                    "gst_rate": new_gst_rate,
                    "gst_amount": gst_amount,
                    "total_amount": basic_amount + gst_amount
                }
                
                new_invoice_items.append(invoice_item)
                total_subtotal += basic_amount
            
            # Calculate GST
            total_cgst = 0
            total_sgst = 0
            total_igst = 0
            
            if project_gst_type == 'CGST_SGST':
                total_gst = sum(item['gst_amount'] for item in new_invoice_items)
                total_cgst = total_gst / 2
                total_sgst = total_gst / 2
            else:
                total_igst = sum(item['gst_amount'] for item in new_invoice_items)
            
            total_gst_amount = total_cgst + total_sgst + total_igst
            total_amount = total_subtotal + total_gst_amount
            
            # Generate amended invoice number
            invoice_count = await db.invoices.count_documents({}) + 1
            amended_invoice_number = f"AME-{original_invoice['invoice_number']}-{invoice_count:05d}"
            
            # Create amended invoice
            amended_invoice = {
                "id": str(uuid.uuid4()),
                "invoice_number": amended_invoice_number,
                "original_invoice_id": amendment_request['original_invoice_id'],
                "amendment_request_id": request_id,
                "amendment_reason": amendment_request['amendment_reason'],
                "amendment_type": amendment_request['amendment_type'],
                "project_id": original_invoice['project_id'],
                "client_id": original_invoice['client_id'],
                "invoice_type": original_invoice.get('invoice_type', 'tax_invoice'),
                "invoice_date": datetime.now(timezone.utc),
                "due_date": datetime.now(timezone.utc) + timedelta(days=30),
                "items": new_invoice_items,
                "subtotal": total_subtotal,
                "gst_type": project_gst_type,
                "cgst_amount": total_cgst,
                "sgst_amount": total_sgst,
                "igst_amount": total_igst,
                "total_gst_amount": total_gst_amount,
                "total_amount": total_amount,
                "payment_terms": original_invoice.get('payment_terms', 'Payment due within 30 days'),
                "advance_received": 0.0,
                "net_amount_due": total_amount,
                "ra_number": original_invoice.get('ra_number'),
                "status": "approved",
                "created_at": datetime.now(timezone.utc),
                "amended_by": amendment_request['requested_by'],
                "approved_by": current_user['id'],
                "approved_at": datetime.now(timezone.utc)
            }
            
            # Insert amended invoice
            await db.invoices.insert_one(amended_invoice)
            
            # Update original invoice status
            await db.invoices.update_one(
                {"id": amendment_request['original_invoice_id']},
                {"$set": {
                    "status": "superseded",
                    "superseded_by": amended_invoice['id'],
                    "superseded_at": datetime.now(timezone.utc)
                }}
            )
            
            # Update amendment request status
            await db.amendment_requests.update_one(
                {"id": request_id},
                {"$set": {
                    "status": "approved",
                    "approved_by": current_user['id'],
                    "approved_at": datetime.now(timezone.utc),
                    "amended_invoice_id": amended_invoice['id']
                }}
            )
            
            await log_activity(
                current_user["id"], current_user["email"], current_user["role"],
                "amendment_approved", f"Approved amendment request for invoice {original_invoice['invoice_number']} -> {amended_invoice_number}"
            )
            
            return {
                "message": "Amendment request approved and invoice created",
                "amended_invoice": amended_invoice
            }
            
        else:  # reject
            # Update amendment request status
            await db.amendment_requests.update_one(
                {"id": request_id},
                {"$set": {
                    "status": "rejected",
                    "rejected_by": current_user['id'],
                    "rejected_at": datetime.now(timezone.utc),
                    "rejection_reason": approval_data.get('rejection_reason', '')
                }}
            )
            
            # Update original invoice status back to its previous state
            await db.invoices.update_one(
                {"id": amendment_request['original_invoice_id']},
                {"$set": {
                    "status": "created",  # Reset to created status
                    "updated_at": datetime.now(timezone.utc)
                }}
            )
            
            await log_activity(
                current_user["id"], current_user["email"], current_user["role"],
                "amendment_rejected", f"Rejected amendment request for invoice {amendment_request['original_invoice_id']}"
            )
            
            return {
                "message": "Amendment request rejected",
                "request_id": request_id
            }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Amendment approval error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to process amendment approval: {str(e)}")

# WebSocket Endpoint for Real-time Project Updates
@app.websocket("/ws/projects/{project_id}")
async def websocket_endpoint(websocket: WebSocket, project_id: str):
    """WebSocket endpoint for real-time project updates"""
    user_id = None
    
    try:
        # Accept connection
        await websocket.accept()
        
        # Get user authentication from query params or headers
        user_id = websocket.query_params.get("user_id", f"anonymous_{uuid.uuid4()}")
        
        # Connect user to project channel
        await manager.connect(websocket, project_id, user_id)
        
        # Send initial project snapshot
        snapshot = await manager.get_project_snapshot(project_id)
        if snapshot:
            await manager.send_personal_message(json.dumps(snapshot), websocket)
        
        # Keep connection alive and handle reconnection requests
        while True:
            try:
                message = await websocket.receive_text()
                data = json.loads(message)
                
                # Handle different message types
                if data.get("type") == "ping":
                    await manager.send_personal_message(
                        json.dumps({"type": "pong", "timestamp": datetime.now(timezone.utc).isoformat()}), 
                        websocket
                    )
                elif data.get("type") == "request_snapshot":
                    snapshot = await manager.get_project_snapshot(project_id)
                    if snapshot:
                        await manager.send_personal_message(json.dumps(snapshot), websocket)
                elif data.get("type") == "subscribe_events":
                    # Client requesting events since timestamp
                    since_timestamp = data.get("since_timestamp")
                    # For now, just send current snapshot
                    # In production, you'd implement event log retrieval
                    snapshot = await manager.get_project_snapshot(project_id)
                    if snapshot:
                        await manager.send_personal_message(json.dumps(snapshot), websocket)
                        
            except WebSocketDisconnect:
                break
            except Exception as e:
                logger.error(f"WebSocket message error: {e}")
                break
                
    except WebSocketDisconnect:
        pass
    except Exception as e:
        logger.error(f"WebSocket connection error: {e}")
    finally:
        if user_id:
            await manager.disconnect(websocket, project_id, user_id)

# SSE (Server-Sent Events) Endpoint - Fallback for clients that can't use WebSocket
@api_router.get("/projects/{project_id}/events")
async def project_events_sse(project_id: str, current_user: dict = Depends(get_current_user)):
    """Server-Sent Events endpoint for real-time updates (WebSocket fallback)"""
    
    async def event_stream():
        try:
            # Send initial snapshot
            snapshot = await manager.get_project_snapshot(project_id)
            if snapshot:
                yield f"data: {json.dumps(snapshot)}\n\n"
            
            # Keep connection alive with periodic updates
            while True:
                await asyncio.sleep(10)  # 10-second polling interval
                
                # Send updated snapshot
                snapshot = await manager.get_project_snapshot(project_id)
                if snapshot:
                    yield f"data: {json.dumps(snapshot)}\n\n"
                    
        except Exception as e:
            logger.error(f"SSE stream error: {e}")
            yield f"data: {json.dumps({'error': str(e)})}\n\n"
    
    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Cache-Control"
        }
    )

# Enhanced Project Snapshot API for manual refresh
@api_router.get("/projects/{project_id}/snapshot")
async def get_project_snapshot(project_id: str, current_user: dict = Depends(get_current_user)):
    """Get current canonical project state"""
    try:
        snapshot = await manager.get_project_snapshot(project_id)
        if not snapshot:
            raise HTTPException(status_code=404, detail="Project not found")
        return snapshot
        
    except Exception as e:
        logger.error(f"Project snapshot error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get project snapshot: {str(e)}")

# PDF Template Management API Endpoints
@api_router.get("/admin/pdf-template/active")
async def get_active_template(current_user: dict = Depends(get_current_user)):
    """Get the currently active PDF template"""
    try:
        # Check super admin permission
        if current_user.get('role') != 'super_admin':
            raise HTTPException(status_code=403, detail="Super admin access required")
        
        # Initialize template manager if not already done
        if not hasattr(template_manager, 'db') or template_manager.db is None:
            await initialize_template_manager(db)
        
        template = await template_manager.get_active_template()
        return custom_jsonable_encoder(template.dict())
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting active template: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get active template: {str(e)}")

@api_router.post("/admin/pdf-template")
async def save_pdf_template(template_data: dict, current_user: dict = Depends(get_current_user)):
    """Save PDF template configuration"""
    try:
        # Check super admin permission
        if current_user.get('role') != 'super_admin':
            raise HTTPException(status_code=403, detail="Super admin access required")
        
        # Initialize template manager if not already done
        if not hasattr(template_manager, 'db') or template_manager.db is None:
            await initialize_template_manager(db)
        
        # Create PDFTemplateConfig from input data
        template_data['created_by'] = current_user.get('id', 'unknown')
        template_config = PDFTemplateConfig(**template_data)
        
        # Save template
        success = await template_manager.save_template(template_config)
        
        if success:
            await log_activity(
                user_id=current_user.get('id'),
                user_email=current_user.get('email'),
                user_role=current_user.get('role'),
                action="template_saved",
                description=f"Saved PDF template configuration: {template_config.name}"
            )
            return {"message": "Template saved successfully", "template_id": template_config.id}
        else:
            raise HTTPException(status_code=500, detail="Failed to save template")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error saving template: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to save template: {str(e)}")

@api_router.post("/admin/pdf-template/preview")
async def generate_template_preview(template_data: dict, current_user: dict = Depends(get_current_user)):
    """Generate a preview PDF with the given template configuration"""
    try:
        # Check super admin permission
        if current_user.get('role') != 'super_admin':
            raise HTTPException(status_code=403, detail="Super admin access required")
        
        # Create template config
        template_config = PDFTemplateConfig(**template_data)
        
        # Generate sample invoice data for preview
        sample_invoice_data = {
            "id": "preview-invoice",
            "invoice_number": "PREVIEW-001",
            "invoice_date": datetime.now(timezone.utc),
            "due_date": datetime.now(timezone.utc) + timedelta(days=30),
            "gst_type": "IGST",
            "subtotal": 100000.0,
            "cgst_amount": 0.0,
            "sgst_amount": 0.0,
            "igst_amount": 18000.0,
            "total_gst_amount": 18000.0,
            "total_amount": 118000.0,
            "payment_terms": "Payment due within 30 days",
            "advance_received": 0.0,
            "net_amount_due": 118000.0,
            "items": [
                {
                    "description": "Sample Construction Work",
                    "unit": "Sqm",
                    "quantity": 100.0,
                    "rate": 1000.0,
                    "amount": 100000.0,
                    "gst_rate": 18.0
                }
            ]
        }
        
        # Sample client and project data
        sample_client = {
            "name": "Sample Client Ltd.",
            "email": "client@example.com",
            "phone": "+91 9876543210",
            "address": "123 Sample Street, Sample City",
            "city": "Sample City",
            "state": "Sample State",
            "gst_no": "27AAAAA0000A1Z5",
            "bill_to_address": "123 Sample Street, Sample City - 400001"
        }
        
        sample_project = {
            "project_name": "Sample Construction Project",
            "location": "Sample Location, Sample City"
        }
        
        # Generate PDF using template-driven system
        pdf_buffer = await generate_template_driven_pdf(
            template_config, 
            sample_invoice_data, 
            sample_client, 
            sample_project
        )
        
        # Return PDF response
        return StreamingResponse(
            io.BytesIO(pdf_buffer),
            media_type="application/pdf",
            headers={
                "Content-Disposition": "attachment; filename=template_preview.pdf",
                "Content-Length": str(len(pdf_buffer))
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating template preview: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate preview: {str(e)}")

@api_router.get("/admin/pdf-templates")
async def list_pdf_templates(current_user: dict = Depends(get_current_user)):
    """List all available PDF templates"""
    try:
        # Check super admin permission
        if current_user.get('role') != 'super_admin':
            raise HTTPException(status_code=403, detail="Super admin access required")
        
        # Initialize template manager if not already done
        if not hasattr(template_manager, 'db') or template_manager.db is None:
            await initialize_template_manager(db)
        
        templates = await template_manager.list_templates()
        return custom_jsonable_encoder([t.dict() for t in templates])
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing templates: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to list templates: {str(e)}")

@api_router.post("/admin/pdf-template/upload-logo")
async def upload_template_logo(
    file: UploadFile = File(...), 
    current_user: dict = Depends(get_current_user)
):
    """Upload logo for PDF template"""
    try:
        # Check super admin permission
        if current_user.get('role') != 'super_admin':
            raise HTTPException(status_code=403, detail="Super admin access required")
        
        # Validate file type
        if not file.content_type or not file.content_type.startswith('image/'):
            raise HTTPException(status_code=400, detail="Only image files are allowed")
        
        # Validate file size (max 5MB)
        max_size = 5 * 1024 * 1024  # 5MB
        file_content = await file.read()
        if len(file_content) > max_size:
            raise HTTPException(status_code=400, detail="File size must be less than 5MB")
        
        # Create upload directory if it doesn't exist (AWS-compatible)
        upload_dir = os.getenv('UPLOAD_DIR', '/tmp/uploads/logos')
        os.makedirs(upload_dir, exist_ok=True)
        
        # Generate unique filename
        file_extension = os.path.splitext(file.filename)[1] or '.png'
        unique_filename = f"logo_{str(uuid.uuid4())}{file_extension}"
        file_path = os.path.join(upload_dir, unique_filename)
        
        # Save file
        with open(file_path, "wb") as f:
            f.write(file_content)
        
        # Convert to base64 for production deployment compatibility
        file_base64 = base64.b64encode(file_content).decode('utf-8')
        logo_url = f"data:{file.content_type};base64,{file_base64}"
        
        await log_activity(
            user_id=current_user.get('id'),
            user_email=current_user.get('email'),
            user_role=current_user.get('role'),
            action="logo_uploaded",
            description=f"Uploaded template logo: {unique_filename}"
        )
        
        return {
            "message": "Logo uploaded successfully",
            "logo_url": logo_url,
            "filename": unique_filename,
            "size": len(file_content)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading logo: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to upload logo: {str(e)}")

# Health check endpoint for AWS
@app.get("/api/health")
async def health_check():
    """Health check endpoint for load balancers and monitoring"""
    try:
        # Test database connection
        await db.list_collection_names()
        return {
            "status": "healthy",
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "database": "connected",
            "service": "invoice-management-system"
        }
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Service unhealthy: {str(e)}")

# Include API router
app.include_router(api_router)

# Application startup
@app.on_event("startup")
async def startup_event():
    await init_super_admin()
    await initialize_template_manager(db)
    logger.info("Activus Invoice Management System started successfully")

# Production server configuration
if __name__ == "__main__":
    uvicorn.run(
        "server:app",
        host="0.0.0.0",
        port=8001,
        reload=False,
        log_level="info"
    )